"""PM-ready Alphabet model, workbook, and memo.

This module is intentionally narrow.  The generic two-company engine remains
useful for broad research, but Alphabet's initiation question needs a dated
stub, a quarterly placed-in-service depreciation schedule, a point-in-time
share bridge, and an explicit trade-readiness gate.  Those controls live here
and drive both reader-facing artifacts.
"""
from __future__ import annotations

import json
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DATA = Path(__file__).resolve().parent / "data"
OUT_JSON = DATA / "alphabet_pm.json"
VALUATION_DATE = date(2026, 8, 7)
FINANCIAL_CUTOFF = date(2026, 6, 30)
YEARS = list(range(2026, 2032))

SEC_10Q = "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000071/goog-20260630.htm"
SEC_RELEASE = "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000066/googexhibit991q22026.htm"
SEC_10K = "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/goog-20251231.htm"
Q2_CALL = "https://alphabet2025ir.q4web.com/investor/events/event-details/2026/2026-Q2-Earnings-Call-2026-GgTAq7Is0z/default.aspx"
TREASURY = ("https://home.treasury.gov/resource-center/data-chart-center/interest-rates/"
            "TextView?type=daily_treasury_yield_curve&field_tdr_date_value=2026")


INPUTS = {
    "spot": 354.24,
    "spot_as_of": "2026-08-07 15:59:59 ET",
    "spot_source": "Robinhood last regular-hours trade",
    "adv_shares_30d": 31_119_520.5407,
    "h1_revenue": 229.692e9,
    "h1_ebit": 80.466e9,
    "h1_ppe_depreciation": 13.586e9,
    "h1_intangible_amortization": 0.545e9,
    "h1_da": 14.131e9,
    "q2_ppe_depreciation": 7.104e9,
    "q2_intangible_amortization": 0.367e9,
    "q2_da": 7.471e9,
    "h1_capex": 80.598e9,
    "h1_cfo": 84.859e9,
    "fy26_revenue": 498.422507321e9,
    "growth": [None, 0.1583, 0.1359, 0.1204, 0.1083, 0.0953],
    "ebitda_margin": [0.385, 0.392, 0.398, 0.402, 0.405, 0.407],
    # FY27 steps up 15%, consistent with management's "increase significantly"
    # direction.  Every year after FY26 is explicitly an analyst assumption.
    "capex": [200e9, 230e9, 225e9, 205e9, 185e9, 170e9],
    "tax_rate": [0.185] * 6,
    "nwc_pct_revenue": [0.015] * 6,
    "terminal_growth": 0.030,
    "terminal_roic": 0.180,
    "risk_free": 0.0465,
    "beta": 0.95,
    "erp": 0.0450,
    "pretax_debt_cost": 0.0480,
    "debt": 100.164e9,
    "leases": 18.037e9,
    "liquid_cash_and_securities": 162.474e9,
    "restricted_marketable_equity": 94.126e9,
    "restricted_marketable_haircut": 0.25,
    "nonmarketable_equity": 124.259e9,
    "nonmarketable_haircut": 0.50,
    # Point-in-time common shares on the 10-Q cover, plus ordinary diluted
    # instruments and minimum mandatory-convertible dilution.  Preferred is
    # therefore not deducted again in the equity bridge.
    "common_shares": 12.230e9,
    "q2_basic_weighted_shares": 12.151e9,
    "q2_diluted_weighted_shares": 12.309e9,
    "ordinary_dilution": 142e6,
    "series_a_pref_shares": 9.625e6,
    "series_b_pref_shares": 9.625e6,
    "series_a_min_conversion": 2.2520,
    "series_b_min_conversion": 2.2740,
    # Capex placement assumptions.  These are analyst estimates, not disclosed
    # asset mix.  The opening quarterly D&A is anchored to reported Q2 D&A.
    "asset_buckets": [
        {"name": "Servers / machines", "mix": 0.60, "life_years": 6.0, "lag_quarters": 1},
        {"name": "Network equipment", "mix": 0.20, "life_years": 6.0, "lag_quarters": 2},
        {"name": "Buildings / improvements", "mix": 0.20, "life_years": 25.0, "lag_quarters": 4},
    ],
    "street_fy27_eps": 15.00595,
    "street_fy27_eps_low": 13.92609,
    "street_fy27_eps_high": 16.67667,
    "street_fy27_n": 42,
    "street_fy27_ebitda": 225.057923702e9,
    "street_fy27_ebit": 197.033766501e9,
    "consensus_freeze": "FMP feed retrieved 2026-08-07; independent $14.20-$14.68 range unresolved",
}

SCENARIOS = [
    {
        "key": "bear", "label": "Stock bear / short works", "probability": 0.25,
        "da_factor": 1.10, "ebitda_offset": 0.40, "multiple": 20.0,
        "basis": "D&A reaches or exceeds path; EBITDA offsets <50%; multiple compresses.",
    },
    {
        "key": "base", "label": "Base / wait", "probability": 0.50,
        "da_factor": 1.00, "ebitda_offset": 0.75, "multiple": 23.0,
        "basis": "D&A rises, but most of the charge is absorbed; catalyst is insufficient.",
    },
    {
        "key": "bull", "label": "Stock bull / short loses", "probability": 0.25,
        "da_factor": 0.85, "ebitda_offset": 1.25, "multiple": 26.0,
        "basis": "Commissioning is slower and EBITDA more than offsets the D&A surprise.",
    },
]

IMPLEMENTATION_GATES = [
    ("Catalyst / revision", "NOT CLEARED", "No observed FY2027 EPS cuts from reported D&A."),
    ("Valuation", "NOT CLEARED", "Scenario target needs both an EPS cut and a lower multiple."),
    ("Borrow / carry", "MISSING", "Locate, availability, fee/rebate, utilization, recall and dividend carry."),
    ("Crowding / squeeze", "MISSING", "Short interest, days-to-cover, ownership and buyback execution."),
    ("Options", "MISSING", "Expiry, strikes, bid/ask, IV/skew, Greeks, OI, theta and max loss."),
    ("Hedge", "MISSING", "Alpha definition, factor map, hedge candidate/ratio and basis-risk failure case."),
    ("ADV / exit", "PRELIMINARY", "30-day GOOGL ADV observed; portfolio size and stressed exit plan absent."),
]

ACTION_RULES = [
    ("Initiate starter only if", "Actual D&A reaches/exceeds the frozen path; EBITDA offsets <50% of the surprise; FY2027 consensus EPS falls at least 5% within 10 trading days; net downside remains at least 20%; every implementation gate clears."),
    ("Add", "A second consecutive print confirms the bridge, cumulative FY2027 EPS cuts reach at least 10%, and downside remains at least 2x defined stress loss."),
    ("Stand down / cover", "Two prints show D&A below model, EBITDA fully offsets the charge, FY2027 EPS is flat/up through Q1 2027, useful lives or commissioning materially defer D&A, or the catalyst window expires."),
    ("Implementation trim / cover", "Carry consumes more than 25% of expected gross alpha, borrow/recall worsens, squeeze risk changes materially, or the target gap closes."),
    ("Hedge / resize", "The exposure becomes primarily mega-cap technology beta, rates/duration, or the AI factor instead of D&A revision alpha."),
]

SOURCES = [
    ("S1", "Filed fact", "H1 revenue, EBIT, PPE depreciation, intangible amortization, capex, CFO; balance sheet; capital structure", "2026-06-30", "Alphabet Q2 2026 Form 10-Q", "0001652044-26-000071", SEC_10Q),
    ("S2", "Filed fact / guidance", "Q2 results; ATM unsold at June 30", "2026-07-22", "Alphabet Q2 2026 earnings release", "0001652044-26-000066", SEC_RELEASE),
    ("S2A", "Management guidance", "FY2026 capex $195-205bn; 2027 expected to increase significantly; Q2 infrastructure capex mix", "2026-07-22", "Alphabet Q2 2026 earnings call", "management remarks", Q2_CALL),
    ("S2B", "Filed accounting policy", "Servers/network generally 6 years; buildings 7-40 years", "2025-12-31", "Alphabet FY2025 Form 10-K", "0001652044-26-000018", SEC_10K),
    ("S3", "Market", "GOOGL price, 30-day share volume, shares outstanding", "2026-08-07", "Robinhood equity quote/fundamentals", "read-only MCP snapshot", "https://robinhood.com/stocks/GOOGL"),
    ("S4", "Market", "10-year Treasury par yield 4.65%", "2026-08-07", "U.S. Treasury daily yield curve", "10-year par yield", TREASURY),
    ("S5", "Consensus conflict", "FY2027 EPS $15.01 / 42 estimates; range $13.93-$16.68", "2026-08-07", "FMP analyst financial estimates", "repository snapshot", "https://financialmodelingprep.com/developer/docs"),
    ("A1", "Analyst assumption", "Beta 0.95; ERP 4.50%; terminal ROIC 18%", "2026-08-08", "Independent model assumptions", "not a sourced fact", ""),
    ("A2", "Analyst assumption", "Capex asset mix, useful lives, and placement lags", "2026-08-08", "Independent model assumptions", "not disclosed by Alphabet", ""),
    ("A3", "Analyst assumption", "Scenario probabilities, EBITDA offsets, and exit multiples", "2026-08-08", "Illustrative risk frame", "not an actionable trade", ""),
]


def quarter_labels() -> list[str]:
    return [f"Q{q} {y}" for y in range(2026, 2032) for q in range(1, 5)
            if (y > 2026 or q >= 3)]


def annual_revenue() -> list[float]:
    out = [INPUTS["fy26_revenue"]]
    for g in INPUTS["growth"][1:]:
        out.append(out[-1] * (1 + g))
    return out


def quarterly_capex() -> list[float]:
    out = [(INPUTS["capex"][0] - INPUTS["h1_capex"]) / 2] * 2
    for annual in INPUTS["capex"][1:]:
        out.extend([annual / 4] * 4)
    return out


def depreciation_schedule() -> dict:
    labels = quarter_labels()
    capex = quarterly_capex()
    legacy = [INPUTS["q2_da"]] * len(labels)
    contributions: dict[str, list[float]] = {}
    for bucket in INPUTS["asset_buckets"]:
        vals = []
        for q in range(len(labels)):
            total = 0.0
            for vintage, spend in enumerate(capex):
                if q >= vintage + bucket["lag_quarters"]:
                    total += spend * bucket["mix"] / (bucket["life_years"] * 4)
            vals.append(total)
        contributions[bucket["name"]] = vals
    total = [legacy[i] + sum(v[i] for v in contributions.values()) for i in range(len(labels))]
    annual = {y: 0.0 for y in YEARS}
    annual_legacy = {y: 0.0 for y in YEARS}
    annual_forecast = {y: 0.0 for y in YEARS}
    annual[2026] = INPUTS["h1_da"]
    annual_legacy[2026] = INPUTS["h1_da"]
    for i, label in enumerate(labels):
        y = int(label[-4:])
        annual[y] += total[i]
        annual_legacy[y] += legacy[i]
        annual_forecast[y] += total[i] - legacy[i]
    return {"labels": labels, "capex": capex, "legacy": legacy,
            "contributions": contributions, "total": total, "annual": annual,
            "annual_legacy": annual_legacy, "annual_forecast": annual_forecast}


def share_bridge() -> dict:
    ordinary = INPUTS["ordinary_dilution"]
    pref = (INPUTS["series_a_pref_shares"] * INPUTS["series_a_min_conversion"]
            + INPUTS["series_b_pref_shares"] * INPUTS["series_b_min_conversion"])
    total = INPUTS["common_shares"] + ordinary + pref
    return {"common": INPUTS["common_shares"], "ordinary_dilution": ordinary,
            "preferred_if_converted": pref, "total": total,
            "preferred_equity_deduction": 0.0,
            "treatment": "if-converted: dilution included; preferred carrying value not deducted"}


def wacc_build(shares: float) -> dict:
    market_cap = shares * INPUTS["spot"]
    cost_equity = INPUTS["risk_free"] + INPUTS["beta"] * INPUTS["erp"]
    after_tax_debt = INPUTS["pretax_debt_cost"] * (1 - INPUTS["tax_rate"][0])
    equity_weight = market_cap / (market_cap + INPUTS["debt"])
    debt_weight = 1 - equity_weight
    wacc = equity_weight * cost_equity + debt_weight * after_tax_debt
    return {"market_cap": market_cap, "cost_equity": cost_equity,
            "after_tax_debt_cost": after_tax_debt, "equity_weight": equity_weight,
            "debt_weight": debt_weight, "wacc": wacc}


def dcf_model(dep: dict, shares: dict, wacc: dict) -> dict:
    revenue = annual_revenue()
    q = dep["labels"]
    q3, q4 = q.index("Q3 2026"), q.index("Q4 2026")
    stub_days, h2_days = 146, 184
    stub_fraction = stub_days / h2_days
    h1_ebitda = INPUTS["h1_ebit"] + INPUTS["h1_da"]
    h2_ebitda = revenue[0] * INPUTS["ebitda_margin"][0] - h1_ebitda
    stub_da = dep["total"][q3] * (54 / 92) + dep["total"][q4]
    stub_capex = dep["capex"][q3] * (54 / 92) + dep["capex"][q4]
    rows = []
    for i, year in enumerate(YEARS):
        if year == 2026:
            r = (revenue[0] - INPUTS["h1_revenue"]) * stub_fraction
            ebitda = h2_ebitda * stub_fraction
            da = stub_da
            capex = stub_capex
            exponent = 0.1986
            delta_nwc = r * INPUTS["nwc_pct_revenue"][i]
        else:
            r = revenue[i]
            ebitda = r * INPUTS["ebitda_margin"][i]
            da = dep["annual"][year]
            capex = INPUTS["capex"][i]
            exponent = 0.8986 + (year - 2027)
            delta_nwc = (revenue[i] - revenue[i - 1]) * INPUTS["nwc_pct_revenue"][i]
        ebit = ebitda - da
        nopat = ebit * (1 - INPUTS["tax_rate"][i])
        fcff = nopat + da - capex - delta_nwc
        pv = fcff / (1 + wacc["wacc"]) ** exponent
        rows.append({"year": year, "label": "2026 stub" if year == 2026 else str(year),
                     "revenue": r, "ebitda": ebitda, "da": da, "ebit": ebit,
                     "nopat": nopat, "capex": capex, "delta_nwc": delta_nwc,
                     "fcff": fcff, "exponent": exponent, "pv": pv})
    last = rows[-1]
    next_nopat = last["nopat"] * (1 + INPUTS["terminal_growth"])
    reinvestment_rate = INPUTS["terminal_growth"] / INPUTS["terminal_roic"]
    terminal_fcff = next_nopat * (1 - reinvestment_rate)
    terminal_value = terminal_fcff / (wacc["wacc"] - INPUTS["terminal_growth"])
    terminal_exponent = 5.4
    pv_terminal = terminal_value / (1 + wacc["wacc"]) ** terminal_exponent
    h2_revenue = revenue[0] - INPUTS["h1_revenue"]
    h2_ebit = h2_ebitda - (dep["total"][q3] + dep["total"][q4])
    h2_nopat = h2_ebit * (1 - INPUTS["tax_rate"][0])
    h2_fcff = (h2_nopat + dep["total"][q3] + dep["total"][q4]
               - (INPUTS["capex"][0] - INPUTS["h1_capex"])
               - h2_revenue * INPUTS["nwc_pct_revenue"][0])
    interim_fcff = h2_fcff * (38 / 184)
    enterprise_value = sum(r["pv"] for r in rows) + pv_terminal
    restricted_value = INPUTS["restricted_marketable_equity"] * (1 - INPUTS["restricted_marketable_haircut"])
    nonmarketable_value = INPUTS["nonmarketable_equity"] * (1 - INPUTS["nonmarketable_haircut"])
    equity_value = (enterprise_value + INPUTS["liquid_cash_and_securities"] + restricted_value + nonmarketable_value
                    + interim_fcff - INPUTS["debt"] - INPUTS["leases"])
    return {"rows": rows, "terminal": {"next_nopat": next_nopat,
            "reinvestment_rate": reinvestment_rate, "normalized_reinvestment": next_nopat * reinvestment_rate,
            "fcff": terminal_fcff, "value": terminal_value, "exponent": terminal_exponent,
            "pv": pv_terminal}, "estimated_interim_fcff": interim_fcff,
            "enterprise_value": enterprise_value, "restricted_marketable_value": restricted_value,
            "nonmarketable_value": nonmarketable_value, "equity_value": equity_value,
            "value_per_share": equity_value / shares["total"],
            "tv_pct_ev": pv_terminal / enterprise_value,
            "stub_fraction": stub_fraction}


def scenario_model(dep: dict, shares: dict) -> dict:
    street_da = INPUTS["street_fy27_ebitda"] - INPUTS["street_fy27_ebit"]
    model_da = dep["annual"][2027]
    rows = []
    for s in SCENARIOS:
        case_da = model_da * s["da_factor"]
        gross_da_surprise = case_da - street_da
        unoffset = gross_da_surprise * (1 - s["ebitda_offset"])
        eps_delta = -unoffset * (1 - INPUTS["tax_rate"][1]) / shares["total"]
        eps = INPUTS["street_fy27_eps"] + eps_delta
        target = eps * s["multiple"]
        rows.append({**s, "street_da": street_da, "case_da": case_da,
                     "gross_da_surprise": gross_da_surprise, "unoffset_da": unoffset,
                     "eps_delta": eps_delta, "eps": eps, "target": target,
                     "return": target / INPUTS["spot"] - 1})
    ev = sum(r["target"] * r["probability"] for r in rows)
    return {"rows": rows, "expected_value": ev, "expected_return": ev / INPUTS["spot"] - 1}


def checks(dep: dict, shares: dict, wacc: dict, dcf: dict, scenarios: dict) -> list[dict]:
    raw = [
        ("No pre-valuation FCFF in DCF", dcf["rows"][0]["label"] == "2026 stub", "DCF begins 2026-08-08"),
        ("WACC less g >= 150bp", wacc["wacc"] - INPUTS["terminal_growth"] >= 0.015, f"{wacc['wacc'] - INPUTS['terminal_growth']:.2%}"),
        ("Terminal ROIC exceeds g", INPUTS["terminal_roic"] > INPUTS["terminal_growth"], f"{INPUTS['terminal_roic']:.1%} vs {INPUTS['terminal_growth']:.1%}"),
        ("Terminal reinvestment is 0-100%", 0 <= dcf["terminal"]["reinvestment_rate"] <= 1, f"{dcf['terminal']['reinvestment_rate']:.1%}"),
        ("Scenario probabilities sum to 100%", abs(sum(s["probability"] for s in SCENARIOS) - 1) < 1e-9, "100.0%"),
        ("Preferred not double counted", shares["preferred_if_converted"] > 0 and shares["preferred_equity_deduction"] == 0, shares["treatment"]),
        ("FY2026 capex within guidance", 195e9 <= INPUTS["capex"][0] <= 205e9, f"${INPUTS['capex'][0]/1e9:.0f}bn"),
        ("FY2027 capex increases significantly", INPUTS["capex"][1] / INPUTS["capex"][0] - 1 >= 0.10, f"{INPUTS['capex'][1] / INPUTS['capex'][0] - 1:.1%}"),
        ("D&A separated by vintage", dep["annual_forecast"][2027] > 0 and dep["annual_legacy"][2027] > 0, "legacy + forecast shown"),
        ("Current position is zero", True, "WAIT FOR PROOF / 0.0%"),
    ]
    return [{"name": n, "status": "PASS" if ok else "FAIL", "detail": detail} for n, ok, detail in raw]


def build_data(write: bool = True) -> dict:
    dep = depreciation_schedule()
    shares = share_bridge()
    wacc = wacc_build(shares["total"])
    dcf = dcf_model(dep, shares, wacc)
    scenarios = scenario_model(dep, shares)
    data = {"as_of": str(VALUATION_DATE), "financial_cutoff": str(FINANCIAL_CUTOFF),
            "inputs": INPUTS, "depreciation": dep, "share_bridge": shares,
            "wacc": wacc, "dcf": dcf, "scenarios": scenarios,
            "implementation_gates": IMPLEMENTATION_GATES, "action_rules": ACTION_RULES,
            "sources": SOURCES}
    data["checks"] = checks(dep, shares, wacc, dcf, scenarios)
    if write:
        OUT_JSON.write_text(json.dumps(data, indent=2))
    return data


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
NAVY = "13213C"
BLUE = "DCE9F7"
PALE = "F2F4F7"
GREEN = "E7F4EC"
RED = "FCE8EA"
WHITE = "FFFFFF"
MUTED = "667085"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=BLUE)
CALC_FILL = PatternFill("solid", fgColor=PALE)
PASS_FILL = PatternFill("solid", fgColor=GREEN)
FAIL_FILL = PatternFill("solid", fgColor=RED)
HEADER_FONT = Font(color=WHITE, bold=True)
TITLE_FONT = Font(color=NAVY, bold=True, size=15)
SUBTITLE_FONT = Font(color=MUTED, italic=True, size=9)
LABEL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D0D5DD")
BN = '$#,##0.0,,,;[Red]($#,##0.0,,,);-'
USD = '$#,##0.00;[Red]($#,##0.00);-'
PCT = '0.0%;[Red](0.0%);-'
PCT2 = '0.00%;[Red](0.00%);-'


def _sheet(wb: Workbook, name: str, widths: list[float]):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def _title(ws, text: str, note: str | None = None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if note:
        ws["A2"] = note
        ws["A2"].font = SUBTITLE_FONT
    ws.freeze_panes = "B4"


def _header(ws, row: int, labels: list[str]):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row, col, label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", wrap_text=True)


def _input(cell, value, fmt=None):
    cell.value = value
    cell.fill = INPUT_FILL
    cell.font = Font(color="0000FF")
    if fmt:
        cell.number_format = fmt


def _formula(cell, formula, fmt=None, bold=False):
    cell.value = formula
    cell.fill = CALC_FILL
    cell.font = Font(bold=bold)
    if fmt:
        cell.number_format = fmt


def build_workbook(path: Path | None = None) -> Path:
    data = build_data()
    wb = Workbook()
    wb.remove(wb.active)

    cover = _sheet(wb, "Cover", [34, 22, 22, 58])
    _title(cover, "Alphabet (GOOGL) — PM readiness model",
           "Decision first. Blue cells are inputs; gray cells are formulas. Model as of 2026-08-07.")
    rows = [
        (4, "Current stance", "WAIT FOR PROOF / NO POSITION", None),
        (5, "Current size", 0.0, PCT),
        (6, "Calculation integrity", '=IF(Checks!B3="PASS","PASS","FAIL")', None),
        (7, "Trade implementation", "NOT CLEARED", None),
        (9, "Spot", "=Drivers!B14", USD),
        (10, "Primary scenario EV", "=Decision!B8", USD),
        (11, "DCF cross-check", "=Valuation!B34", USD),
        (12, "FY2027 restated EPS", "=Variant!B9", USD),
        (13, "WACC", "=WACC!B15", PCT2),
        (14, "Terminal value / EV", "=Valuation!B31", PCT),
    ]
    for r, label, value, fmt in rows:
        cover.cell(r, 1, label).font = LABEL_FONT
        if isinstance(value, str) and value.startswith("="):
            _formula(cover.cell(r, 2), value, fmt, bold=True)
        else:
            cover.cell(r, 2, value)
            if fmt: cover.cell(r, 2).number_format = fmt
    cover["A17"] = "Selected causal thesis"
    cover["A17"].font = LABEL_FONT
    cover["A18"] = ("AI capex converts into GAAP D&A faster than consensus EPS incorporates it; "
                    "only if EBITDA fails to absorb the charge do quarterly prints force FY2027-28 "
                    "EPS cuts and a multiple reset. DCF is a risk frame, not the catalyst.")
    cover.merge_cells("A18:D20")
    cover["A18"].alignment = Alignment(wrap_text=True, vertical="top")
    cover["A22"] = "Circulation posture"
    cover["A22"].font = LABEL_FONT
    cover["A23"] = ("Shareable as a research-monitoring package. Not approved to risk capital: "
                    "borrow/carry, crowding, options, hedge and observed revision evidence remain open.")
    cover.merge_cells("A23:D25")
    cover["A23"].alignment = Alignment(wrap_text=True, vertical="top")

    src = _sheet(wb, "Sources", [10, 20, 46, 14, 34, 28, 60])
    _title(src, "Source ledger", "Sourced facts and analyst assumptions are deliberately separated.")
    _header(src, 4, ["ID", "Type", "Metric / assumption", "As of", "Document / provider", "Accession / location", "URL"])
    for r, row in enumerate(SOURCES, 5):
        for c, value in enumerate(row, 1):
            src.cell(r, c, value)
            src.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        if row[6]: src.cell(r, 7).hyperlink = row[6]

    drv = _sheet(wb, "Drivers", [36, 16, 16, 16, 16, 16, 16, 52])
    _title(drv, "Operating and valuation drivers", "FY2026 contains reported H1 plus forecast H2; only the remaining stub is discounted.")
    _header(drv, 4, ["Driver"] + [str(y) for y in YEARS] + ["Evidence label / note"])
    driver_rows = {}
    annual = annual_revenue()
    entries = [
        ("Full-year revenue", annual, BN, "2026 = reported H1 + forecast H2; later years formula driven"),
        ("Revenue growth", INPUTS["growth"], PCT, "Analyst assumption after 2026"),
        ("EBITDA margin", INPUTS["ebitda_margin"], PCT, "Analyst assumption"),
        ("Capex", INPUTS["capex"], BN, "FY2026 midpoint of $195-205bn guidance; later years assumption"),
        ("Tax rate", INPUTS["tax_rate"], PCT, "Analyst assumption"),
        ("NWC % revenue", INPUTS["nwc_pct_revenue"], PCT, "Analyst assumption"),
    ]
    for r, (label, values, fmt, note) in enumerate(entries, 5):
        driver_rows[label] = r
        drv.cell(r, 1, label).font = LABEL_FONT
        for i, value in enumerate(values):
            cell = drv.cell(r, 2 + i)
            if label == "Full-year revenue" and i > 0:
                _formula(cell, f"={get_column_letter(1+i)}{r}*(1+{get_column_letter(2+i)}{driver_rows['Revenue growth'] if 'Revenue growth' in driver_rows else r+1})", fmt)
            elif label == "Full-year revenue" and i == 0:
                _input(cell, value, fmt)
            else:
                _input(cell, value, fmt)
        drv.cell(r, 8, note).font = SUBTITLE_FONT
    # Repair revenue formulas now that the growth row is known.
    for i in range(1, 6):
        col, prev = get_column_letter(2+i), get_column_letter(1+i)
        drv.cell(driver_rows["Full-year revenue"], 2+i,
                 f"={prev}{driver_rows['Full-year revenue']}*(1+{col}{driver_rows['Revenue growth']})").number_format = BN
        drv.cell(driver_rows["Full-year revenue"], 2+i).fill = CALC_FILL
    scalar_rows = {}
    scalar_inputs = [
        ("Spot price", INPUTS["spot"], USD, "S3; regular-hours last trade"),
        ("Valuation date", VALUATION_DATE, "yyyy-mm-dd", "S3"),
        ("Financial cutoff", FINANCIAL_CUTOFF, "yyyy-mm-dd", "S1"),
        ("H1 revenue", INPUTS["h1_revenue"], BN, "S1"),
        ("H1 EBIT", INPUTS["h1_ebit"], BN, "S1"),
        ("H1 D&A", INPUTS["h1_da"], BN, "S1"),
        ("H1 capex", INPUTS["h1_capex"], BN, "S1"),
        ("Terminal growth", INPUTS["terminal_growth"], PCT, "A1"),
        ("Terminal ROIC", INPUTS["terminal_roic"], PCT, "A1"),
    ]
    for off, (label, value, fmt, note) in enumerate(scalar_inputs, 14):
        scalar_rows[label] = off
        drv.cell(off, 1, label).font = LABEL_FONT
        _input(drv.cell(off, 2), value, fmt)
        drv.cell(off, 3, note).font = SUBTITLE_FONT

    sc = _sheet(wb, "Scenarios", [34, 14, 14, 16, 16, 18, 18, 50])
    _title(sc, "Illustrative EPS × multiple scenario skew", "Probabilities are analyst assumptions. This is not an approved trade or size.")
    _header(sc, 4, ["Case", "Probability", "D&A factor", "EBITDA offset", "Exit multiple", "FY2027 EPS", "Target", "Basis"])
    for r, s in enumerate(SCENARIOS, 5):
        sc.cell(r, 1, s["label"])
        for c, key, fmt in [(2, "probability", PCT), (3, "da_factor", '0.00x'), (4, "ebitda_offset", PCT), (5, "multiple", '0.0x')]:
            _input(sc.cell(r, c), s[key], fmt)
        # EPS and target are live, fed by the base D&A result and Street inputs.
        sc.cell(r, 6, f"=Drivers!B35-((Variant!B7*C{r}-Variant!B6)*(1-D{r}))*(1-Drivers!C9)/'Share Bridge'!B11")
        sc.cell(r, 6).number_format = USD
        sc.cell(r, 7, f"=F{r}*E{r}").number_format = USD
        sc.cell(r, 8, s["basis"]).alignment = Alignment(wrap_text=True)

    wa = _sheet(wb, "WACC", [38, 20, 18, 56])
    _title(wa, "CAPM and capital-structure build", "Market-value weights; preferred is if-converted and therefore sits in equity, not a separate deduction.")
    _header(wa, 4, ["Input / calculation", "Value", "Evidence", "Note"])
    wacc_lines = [
        ("Risk-free rate", INPUTS["risk_free"], PCT2, "S4", "10-year Treasury par yield, 2026-08-07"),
        ("Levered beta", INPUTS["beta"], '0.00x', "A1", "Analyst assumption; refresh before risk"),
        ("Equity risk premium", INPUTS["erp"], PCT2, "A1", "Analyst assumption; refresh before risk"),
        ("Cost of equity", "=B5+B6*B7", PCT2, "Formula", "CAPM"),
        ("Pretax cost of debt", INPUTS["pretax_debt_cost"], PCT2, "S1", "4.80% weighted-average coupon disclosed for 2026 USD notes"),
        ("Marginal tax rate", INPUTS["tax_rate"][0], PCT, "A1", "Operating valuation assumption"),
        ("After-tax debt cost", "=B9*(1-B10)", PCT2, "Formula", ""),
        ("Equity market value", "=Drivers!B14*'Share Bridge'!B11", BN, "Formula", "Spot × point diluted shares"),
        ("Debt", INPUTS["debt"], BN, "S1", "June 30 balance sheet"),
        ("Equity weight", "=B12/(B12+B13)", PCT, "Formula", ""),
        ("WACC", "=B14*B8+(1-B14)*B11", PCT2, "Formula", "Single base WACC definition"),
    ]
    for r, (label, value, fmt, ev, note) in enumerate(wacc_lines, 5):
        wa.cell(r, 1, label).font = LABEL_FONT if label in ("Cost of equity", "WACC") else Font()
        if isinstance(value, str) and value.startswith("="):
            _formula(wa.cell(r, 2), value, fmt, bold=label == "WACC")
        else:
            _input(wa.cell(r, 2), value, fmt)
        wa.cell(r, 3, ev)
        wa.cell(r, 4, note).font = SUBTITLE_FONT

    sh = _sheet(wb, "Share Bridge", [42, 22, 22, 56])
    _title(sh, "Point-in-time diluted share bridge", "The Q2 weighted-average diluted share count is not used as the valuation denominator.")
    _header(sh, 4, ["Line", "Shares", "Evidence", "Treatment"])
    share_lines = [
        ("Class A/B/C outstanding, 2026-07-15", INPUTS["common_shares"], "S1", "Point-in-time common shares"),
        ("Ordinary dilution (RSUs / contingent shares)", INPUTS["ordinary_dilution"], "S1", "Filed Q2 dilution component"),
        ("Q2 basic weighted-average shares", INPUTS["q2_basic_weighted_shares"], "S1", "Reference only"),
        ("Q2 diluted weighted-average shares", INPUTS["q2_diluted_weighted_shares"], "S1", "Reference only; not valuation denominator"),
        ("Series A preferred, minimum conversion", INPUTS["series_a_pref_shares"] * INPUTS["series_a_min_conversion"], "S1", "If-converted"),
        ("Series B preferred, minimum conversion", INPUTS["series_b_pref_shares"] * INPUTS["series_b_min_conversion"], "S1", "If-converted"),
        ("Point-in-time diluted shares", "=B5+B6+B9+B10", "Formula", "Used in valuation"),
        ("Preferred equity deduction", 0.0, "Formula", "Zero: preferred dilution already included"),
    ]
    for r, (label, value, ev, note) in enumerate(share_lines, 5):
        sh.cell(r, 1, label).font = LABEL_FONT if label == "Point-in-time diluted shares" else Font()
        if isinstance(value, str) and value.startswith("="):
            _formula(sh.cell(r, 2), value, '#,##0,,', bold=label == "Point-in-time diluted shares")
        else:
            _input(sh.cell(r, 2), value, '#,##0,,')
        sh.cell(r, 3, ev)
        sh.cell(r, 4, note).alignment = Alignment(wrap_text=True)

    rev = _sheet(wb, "Revenue", [34, 17, 17, 17, 17, 17, 17, 48])
    _title(rev, "Revenue build and H1/stub split", "2026 full-year revenue is H1 actual plus forecast H2; DCF includes only the remaining 2026 stub.")
    _header(rev, 4, ["Line"] + [str(y) for y in YEARS] + ["Note"])
    rev["A5"] = "Full-year revenue"
    rev["A6"] = "Growth"
    for i in range(6):
        col = get_column_letter(2+i)
        _formula(rev.cell(5, 2+i), f"=Drivers!{col}5", BN, bold=True)
        _formula(rev.cell(6, 2+i), f"=Drivers!{col}6", PCT)
    rev["A8"] = "Reported H1 2026 revenue"
    _formula(rev["B8"], "=Drivers!B17", BN)
    rev["A9"] = "Forecast H2 2026 revenue"
    _formula(rev["B9"], "=B5-B8", BN)
    rev["A10"] = "Remaining stub fraction"
    _formula(rev["B10"], "=146/184", PCT)
    rev["A11"] = "Revenue in DCF stub"
    _formula(rev["B11"], "=B9*B10", BN, bold=True)

    depws = _sheet(wb, "Depreciation", [28] + [13] * len(data["depreciation"]["labels"]) + [44])
    _title(depws, "Quarterly placed-in-service depreciation",
           "Opening D&A is anchored to Q2 reported run-rate. Asset mix, useful lives and lags are explicit analyst assumptions.")
    labels = data["depreciation"]["labels"]
    _header(depws, 4, ["Driver / line"] + labels)
    # Bucket assumptions in A6:D8.
    for r, bucket in enumerate(INPUTS["asset_buckets"], 6):
        depws.cell(r, 1, bucket["name"])
        _input(depws.cell(r, 2), bucket["mix"], PCT)
        _input(depws.cell(r, 3), bucket["life_years"], '0.0')
        _input(depws.cell(r, 4), bucket["lag_quarters"], '0')
    depws["A9"] = "Q2 opening D&A / quarter"
    _input(depws["B9"], INPUTS["q2_da"], BN)
    depws["E6"] = "Mix"
    depws["E7"] = "Life years"
    depws["E8"] = "Lag quarters"
    # Quarterly schedules begin at row 12.
    depws["A12"] = "Quarterly capex"
    depws["A13"] = "Opening / legacy D&A"
    bucket_rows = {b["name"]: 14+i for i, b in enumerate(INPUTS["asset_buckets"])}
    for name, r in bucket_rows.items(): depws.cell(r, 1, f"Forecast-vintage D&A — {name}")
    total_da_row = 17
    depws.cell(total_da_row, 1, "Total D&A").font = LABEL_FONT
    for i, label in enumerate(labels):
        col = get_column_letter(2+i)
        year = int(label[-4:])
        annual_col = get_column_letter(2 + YEARS.index(year))
        if year == 2026:
            _formula(depws.cell(12, 2+i), f"=(Drivers!{annual_col}{driver_rows['Capex']}-Drivers!B20)/2", BN)
        else:
            _formula(depws.cell(12, 2+i), f"=Drivers!{annual_col}{driver_rows['Capex']}/4", BN)
        _formula(depws.cell(13, 2+i), "=$B$9", BN)
        for b_idx, bucket in enumerate(INPUTS["asset_buckets"]):
            terms = []
            for v in range(i + 1):
                vcol = get_column_letter(2+v)
                terms.append(f"IF({i-v}>=$D${6+b_idx},{vcol}$12*$B${6+b_idx}/($C${6+b_idx}*4),0)")
            _formula(depws.cell(bucket_rows[bucket["name"]], 2+i), "=" + "+".join(terms), BN)
        _formula(depws.cell(total_da_row, 2+i), f"={col}13+SUM({col}14:{col}16)", BN, bold=True)
    depws["A20"] = "Annual D&A"
    depws["A21"] = "Opening / legacy"
    depws["A22"] = "Forecast vintages"
    for i, y in enumerate(YEARS):
        col = get_column_letter(2+i)
        qcols = [get_column_letter(2+j) for j, lab in enumerate(labels) if int(lab[-4:]) == y]
        if y == 2026:
            _formula(depws.cell(20, 2+i), f"=Drivers!B19+SUM({qcols[0]}17:{qcols[-1]}17)", BN, bold=True)
            _formula(depws.cell(21, 2+i), f"=Drivers!B19+SUM({qcols[0]}13:{qcols[-1]}13)", BN)
        else:
            _formula(depws.cell(20, 2+i), f"=SUM({qcols[0]}17:{qcols[-1]}17)", BN, bold=True)
            _formula(depws.cell(21, 2+i), f"=SUM({qcols[0]}13:{qcols[-1]}13)", BN)
        _formula(depws.cell(22, 2+i), f"={col}20-{col}21", BN)

    val = _sheet(wb, "Valuation", [36, 17, 17, 17, 17, 17, 17, 54])
    _title(val, "Dated FCFF DCF cross-check", "The primary trade frame is FY2027 EPS × multiple. This DCF excludes reported H1 and normalizes terminal reinvestment.")
    _header(val, 4, ["Line", "2026 stub", "2027", "2028", "2029", "2030", "2031", "Note"])
    line_rows = {name: r for r, name in enumerate([
        "Revenue", "EBITDA", "D&A", "EBIT", "Tax rate", "NOPAT", "Capex", "Change in NWC",
        "FCFF", "Discount exponent", "Discount factor", "PV of FCFF"], 5)}
    for name, r in line_rows.items(): val.cell(r, 1, name).font = LABEL_FONT if name in ("EBIT", "NOPAT", "FCFF", "PV of FCFF") else Font()
    for i, y in enumerate(YEARS):
        col = get_column_letter(2+i)
        if y == 2026:
            formulas = {
                "Revenue": "=Revenue!B11",
                "EBITDA": "=(Drivers!B5*Drivers!B7-(Drivers!B18+Drivers!B19))*Revenue!B10",
                "D&A": "=Depreciation!B17*(54/92)+Depreciation!C17",
                "Capex": "=Depreciation!B12*(54/92)+Depreciation!C12",
                "Change in NWC": f"={col}{line_rows['Revenue']}*Drivers!B10",
                "Discount exponent": "=0.1986",
            }
        else:
            dcol = get_column_letter(2+i)
            formulas = {
                "Revenue": f"=Revenue!{dcol}5",
                "EBITDA": f"={col}{line_rows['Revenue']}*Drivers!{dcol}7",
                "D&A": f"=Depreciation!{dcol}20",
                "Capex": f"=Drivers!{dcol}8",
                "Change in NWC": f"=(Revenue!{dcol}5-Revenue!{get_column_letter(1+i)}5)*Drivers!{dcol}10",
                "Discount exponent": f"={0.8986 + (y-2027):.4f}",
            }
        formulas.update({
            "EBIT": f"={col}{line_rows['EBITDA']}-{col}{line_rows['D&A']}",
            "Tax rate": f"=Drivers!{get_column_letter(2+i)}9",
            "NOPAT": f"={col}{line_rows['EBIT']}*(1-{col}{line_rows['Tax rate']})",
            "FCFF": f"={col}{line_rows['NOPAT']}+{col}{line_rows['D&A']}-{col}{line_rows['Capex']}-{col}{line_rows['Change in NWC']}",
            "Discount factor": f"=1/(1+WACC!$B$15)^{col}{line_rows['Discount exponent']}",
            "PV of FCFF": f"={col}{line_rows['FCFF']}*{col}{line_rows['Discount factor']}",
        })
        fmts = {"Tax rate": PCT, "Discount exponent": '0.0000', "Discount factor": '0.0000'}
        for name, formula in formulas.items():
            _formula(val.cell(line_rows[name], 2+i), formula, fmts.get(name, BN))
    out_rows = {
        "PV explicit": 19, "Next-period NOPAT": 20, "Terminal ROIC": 21,
        "Reinvestment rate": 22, "Normalized reinvestment": 23, "Terminal FCFF": 24,
        "Terminal value": 25, "Terminal exponent": 26, "PV terminal": 27,
        "Enterprise value": 28, "Estimated interim FCFF": 29, "Terminal value / EV": 31,
        "Equity value": 33, "Value per share": 34,
    }
    formulas = {
        "PV explicit": "=SUM(B16:G16)",
        "Next-period NOPAT": "=G10*(1+Drivers!B21)",
        "Terminal ROIC": "=Drivers!B22",
        "Reinvestment rate": "=Drivers!B21/B21",
        "Normalized reinvestment": "=B20*B22",
        "Terminal FCFF": "=B20-B23",
        "Terminal value": "=B24/(WACC!B15-Drivers!B21)",
        "Terminal exponent": "=5.4",
        "PV terminal": "=B25/(1+WACC!B15)^B26",
        "Enterprise value": "=B19+B27",
        "Estimated interim FCFF": "=((Drivers!B5*Drivers!B7-(Drivers!B18+Drivers!B19)-SUM(Depreciation!B17:C17))*(1-Drivers!B9)+SUM(Depreciation!B17:C17)-(Drivers!B8-Drivers!B20)-(Drivers!B5-Drivers!B17)*Drivers!B10)*(38/184)",
        "Terminal value / EV": "=B27/B28",
        "Equity value": "=B28+Drivers!B24+Drivers!B27+Drivers!B30+B29-Drivers!B31-Drivers!B32",
        "Value per share": "=B33/'Share Bridge'!B11",
    }
    # Non-operating assets are separated and haircut visibly; restricted
    # SpaceX and non-marketable holdings are not silently treated as cash.
    bridge_driver_lines = [
        ("Liquid cash / securities", INPUTS["liquid_cash_and_securities"], BN, "S1; excludes restricted SpaceX"),
        ("Restricted marketable equity, gross", INPUTS["restricted_marketable_equity"], BN, "S1; $80.0bn current + $14.126bn non-current"),
        ("Restricted marketable haircut", INPUTS["restricted_marketable_haircut"], PCT, "A1"),
        ("Restricted marketable equity, adjusted", "=B25*(1-B26)", BN, "Formula"),
        ("Non-marketable equity, gross", INPUTS["nonmarketable_equity"], BN, "S1"),
        ("Non-marketable haircut", INPUTS["nonmarketable_haircut"], PCT, "A1"),
        ("Non-marketable equity, adjusted", "=B28*(1-B29)", BN, "Formula"),
        ("Debt", INPUTS["debt"], BN, "S1; total carrying value"),
        ("Leases", INPUTS["leases"], BN, "S1; valuation-policy assumption"),
    ]
    for r, (label, value, fmt, evidence) in enumerate(bridge_driver_lines, 24):
        drv.cell(r, 1, label).font = LABEL_FONT
        if isinstance(value, str) and value.startswith("="):
            _formula(drv.cell(r, 2), value, fmt)
        else:
            _input(drv.cell(r, 2), value, fmt)
        drv.cell(r, 3, evidence).font = SUBTITLE_FONT
    for label, r in out_rows.items():
        val.cell(r, 1, label).font = LABEL_FONT
        fmt = PCT if label in ("Terminal ROIC", "Reinvestment rate", "Terminal value / EV") else ('0.0000' if label == "Terminal exponent" else (USD if label == "Value per share" else BN))
        _formula(val.cell(r, 2), formulas[label], fmt, bold=label in ("Enterprise value", "Equity value", "Value per share"))

    var = _sheet(wb, "Variant", [38, 20, 20, 20, 22, 56])
    _title(var, "Consensus and EPS-revision monitor", "The FMP EPS feed is a frozen aggregator snapshot. The independent $14.20-$14.68 range conflicts and remains unresolved.")
    _header(var, 4, ["FY2027 bridge", "Value", "Evidence", "Status", "Formula role", "Comment"])
    variant_lines = [
        ("Street FY2027 EPS", INPUTS["street_fy27_eps"], "S5", "SOURCE CONFLICT", "Input", INPUTS["consensus_freeze"]),
        ("Street FY2027 implied D&A", "=Drivers!B36-Drivers!B37", "S5", "AGGREGATOR", "EBITDA - EBIT", "Not a broker-level D&A ledger"),
        ("Our FY2027 D&A", "=Depreciation!C20", "A2", "MODEL", "Quarterly schedule", "Legacy and forecast vintages shown separately"),
        ("Gross D&A difference", "=B7-B6", "Formula", "MODEL", "Our D&A - Street implied", "Before EBITDA offset"),
        ("Base restated FY2027 EPS", "=B5-(B8*(1-Scenarios!D6))*(1-Drivers!C9)/'Share Bridge'!B11", "Formula", "ILLUSTRATIVE", "After base EBITDA offset", "Not a forecast until quarterly consensus is sourced"),
    ]
    # Put research-only consensus inputs below the valuation bridge.
    for r, (label, value, fmt) in enumerate([("Street FY2027 EPS", INPUTS["street_fy27_eps"], USD), ("Street FY2027 EBITDA", INPUTS["street_fy27_ebitda"], BN), ("Street FY2027 EBIT", INPUTS["street_fy27_ebit"], BN)], 35):
        drv.cell(r, 1, label)
        _input(drv.cell(r, 2), value, fmt)
        drv.cell(r, 3, "S5 — source conflict; research use only").font = SUBTITLE_FONT
    for r, row in enumerate(variant_lines, 5):
        var.cell(r, 1, row[0])
        if isinstance(row[1], str) and row[1].startswith("="):
            _formula(var.cell(r, 2), row[1], USD if "EPS" in row[0] else BN, bold=r == 9)
        else:
            _input(var.cell(r, 2), row[1], USD)
        for c, value in enumerate(row[2:], 3):
            var.cell(r, c, value)
            var.cell(r, c).alignment = Alignment(wrap_text=True)
    var["A12"] = "Quarterly catalyst bridge"
    var["A12"].font = TITLE_FONT
    _header(var, 13, ["Print", "Model D&A", "Frozen consensus D&A", "Actual D&A", "EBITDA offset", "Result"])
    q_map = {"Q3 2026": "B", "Q4 2026": "C", "Q1 2027": "D"}
    for r, (label, col) in enumerate(q_map.items(), 14):
        var.cell(r, 1, label)
        _formula(var.cell(r, 2), f"=Depreciation!{col}17", BN)
        var.cell(r, 3, "MISSING")
        var.cell(r, 4, "PENDING")
        var.cell(r, 5, "PENDING")
        var.cell(r, 6, "PENDING")
    var["A18"] = "Required pre-print fields: quarterly EBITDA, EBIT, implied D&A, tax, below-line income, diluted shares, EPS, and 30/60/90-day revision history."
    var.merge_cells("A18:F19")
    var["A18"].alignment = Alignment(wrap_text=True, vertical="top")

    sen = _sheet(wb, "Sensitivities", [24] + [15] * 7)
    _title(sen, "DCF sensitivity — WACC × terminal growth", "Appendix risk frame. Center cell uses the base WACC and 3.0% terminal growth.")
    growths = [0.02, 0.025, 0.03, 0.035, 0.04, 0.045, 0.05]
    waccs = [data["wacc"]["wacc"] + x for x in (-0.015, -0.01, -0.005, 0, 0.005, 0.01, 0.015)]
    sen["A4"] = "WACC / g"
    for j, g in enumerate(growths, 2): _input(sen.cell(4, j), g, PCT)
    for i, w in enumerate(waccs, 5):
        _input(sen.cell(i, 1), w, PCT)
        for j in range(2, 9):
            # Revalue explicit FCFF and normalized terminal FCFF with selected WACC/g.
            terms = "+".join(f"Valuation!{get_column_letter(2+k)}13/(1+$A{i})^Valuation!{get_column_letter(2+k)}14" for k in range(6))
            f = (f"=({terms}+((Valuation!G10*(1+{get_column_letter(j)}$4)*(1-{get_column_letter(j)}$4/Drivers!$B$22))/"
                 f"($A{i}-{get_column_letter(j)}$4))/(1+$A{i})^Valuation!$B$26+Drivers!$B$24+Drivers!$B$27+Drivers!$B$30+Valuation!$B$29-Drivers!$B$31-Drivers!$B$32)/'Share Bridge'!$B$11")
            _formula(sen.cell(i, j), f, USD)

    dec = _sheet(wb, "Decision", [38, 20, 18, 18, 56])
    _title(dec, "Decision and implementation gate", "No percentage recommendation without borrow, hedge, crowding and portfolio constraints.")
    _header(dec, 4, ["Case", "Target", "Probability", "vs spot", "Basis"])
    for r, srow in enumerate(range(5, 8), 5):
        dec.cell(r, 1, f"=Scenarios!A{srow}")
        _formula(dec.cell(r, 2), f"=Scenarios!G{srow}", USD)
        _formula(dec.cell(r, 3), f"=Scenarios!B{srow}", PCT)
        _formula(dec.cell(r, 4), f"=B{r}/Drivers!B14-1", PCT)
        dec.cell(r, 5, f"=Scenarios!H{srow}")
    dec["A8"] = "Probability-weighted value"
    dec["A8"].font = LABEL_FONT
    _formula(dec["B8"], "=SUMPRODUCT(B5:B7,C5:C7)", USD, bold=True)
    _formula(dec["C8"], "=SUM(C5:C7)", PCT)
    _formula(dec["D8"], "=B8/Drivers!B14-1", PCT)
    dec["A10"] = "Current stance"
    dec["B10"] = "WAIT FOR PROOF / NO POSITION"
    dec["A11"] = "Current size"
    dec["B11"] = 0.0
    dec["B11"].number_format = PCT
    dec["A13"] = "Implementation gates"
    dec["A13"].font = TITLE_FONT
    _header(dec, 14, ["Gate", "Status", "Missing evidence", "", ""])
    for r, gate in enumerate(IMPLEMENTATION_GATES, 15):
        dec.cell(r, 1, gate[0])
        dec.cell(r, 2, gate[1])
        dec.cell(r, 3, gate[2])
        dec.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        dec.cell(r, 3).alignment = Alignment(wrap_text=True)
    dec["A24"] = "Conditional action rules"
    dec["A24"].font = TITLE_FONT
    for r, rule in enumerate(ACTION_RULES, 25):
        dec.cell(r, 1, rule[0]).font = LABEL_FONT
        dec.cell(r, 2, rule[1])
        dec.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        dec.cell(r, 2).alignment = Alignment(wrap_text=True)

    chk = _sheet(wb, "Checks", [42, 18, 66])
    _title(chk, "Blocking checks", "A FAIL blocks circulation. Trade implementation can remain open while calculation integrity passes.")
    chk["A3"] = "Calculation integrity"
    chk["B3"] = '=IF(COUNTIF(B6:B15,"FAIL")=0,"PASS","FAIL")'
    chk["B3"].fill = PASS_FILL
    _header(chk, 5, ["Check", "Status", "Detail"])
    for r, item in enumerate(data["checks"], 6):
        chk.cell(r, 1, item["name"])
        chk.cell(r, 2, item["status"])
        chk.cell(r, 2).fill = PASS_FILL if item["status"] == "PASS" else FAIL_FILL
        chk.cell(r, 3, item["detail"]).alignment = Alignment(wrap_text=True)
    chk["A18"] = "Warnings"
    chk["A18"].font = TITLE_FONT
    warnings = [
        f"Terminal value is {data['dcf']['tv_pct_ev']:.0%} of EV; DCF is a sensitivity, not the initiation thesis.",
        "Broker-level quarterly D&A consensus and revision history are missing.",
        "The $15.01 FMP FY2027 EPS conflicts with an independently cited $14.20-$14.68 range; neither is treated as canonical.",
        "Opening-vintage D&A is held at the Q2 run-rate through the explicit horizon because remaining lives by vintage are not disclosed.",
    ]
    for r, warning in enumerate(warnings, 19): chk.cell(r, 1, warning)

    notes = _sheet(wb, "Notes", [110])
    _title(notes, "Method, limits, and circulation note")
    note_lines = [
        "DECISION. WAIT FOR PROOF / NO POSITION. The package is ready for PM review as a monitoring thesis, not for capital deployment.",
        "CAUSAL CHAIN. Capex -> placed in service -> GAAP D&A -> unoffset EBIT/EPS miss -> consensus revisions -> multiple response. A D&A rise without a net EPS cut is not confirmation.",
        "VALUATION DATE. June 30 balance-sheet facts are bridged to August 7 with an explicitly estimated July 1-August 7 FCFF. The DCF excludes H1 and discounts only August 8 onward.",
        "TERMINAL VALUE. Next-period NOPAT is reduced by growth-required reinvestment (g / terminal ROIC) before the Gordon calculation.",
        "SHARES. Point-in-time common shares plus ordinary dilution and minimum mandatory-convertible dilution. Preferred carrying value is not deducted again.",
        "CONSENSUS. FMP is an aggregator snapshot, not a broker-level ledger. The independent EPS range conflict remains open; the variant is illustrative and cannot clear the catalyst gate.",
        "IMPLEMENTATION. Borrow/carry, crowding/squeeze, options and hedge data are missing. Underlying liquidity alone cannot set size.",
        "This is analytical research on public information, not investment advice, a recommendation, or a solicitation.",
    ]
    for r, line in enumerate(note_lines, 4):
        notes.cell(r, 1, line).alignment = Alignment(wrap_text=True, vertical="top")
        notes.row_dimensions[r].height = 42

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions if ws.title == "Sources" else None
        for row in ws.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.vertical = alignment.vertical or "top"
                cell.alignment = alignment
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    target = path or ROOT / "alphabet" / "model.xlsx"
    wb.save(target)
    return target


# ---------------------------------------------------------------------------
# Memo
# ---------------------------------------------------------------------------
CSS = """
:root{--navy:#13213c;--ink:#172033;--muted:#667085;--line:#d0d5dd;--blue:#2864dc;--pale:#f2f4f7;--red:#b42318;--green:#067647;--warn:#6938ef}
*{box-sizing:border-box}body{margin:0;background:#eef2f6;color:var(--ink);font:15px/1.58 Inter,Aptos,system-ui,sans-serif}.page{max-width:1120px;margin:28px auto;background:white;box-shadow:0 18px 55px rgba(19,33,60,.12)}header{background:linear-gradient(135deg,#13213c,#233b68);color:white;padding:52px 62px}header h1{font-size:44px;line-height:1.05;margin:12px 0}.eyebrow{text-transform:uppercase;letter-spacing:.12em;color:#b9cdf9;font-size:12px;font-weight:800}.dek{font-size:18px;color:#e6ecfa;max-width:880px}.chips{display:flex;gap:9px;flex-wrap:wrap;margin-top:24px}.chip{border:1px solid #ffffff44;border-radius:999px;padding:6px 11px;font-size:12px;font-weight:700}main{padding:44px 62px 58px}h2{font-size:24px;margin:38px 0 12px;border-top:1px solid var(--line);padding-top:28px}h3{font-size:17px;margin:24px 0 8px}.callout{border-left:5px solid var(--blue);background:#edf3ff;padding:18px 20px;margin:18px 0}.warn{border-left-color:var(--warn);background:#f4f3ff}.stop{border-left-color:var(--red);background:#fff1f0}.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.card{border:1px solid var(--line);border-radius:12px;padding:16px}.metric{font-size:24px;font-weight:800}.sub{color:var(--muted);font-size:13px}table{width:100%;border-collapse:collapse;margin:12px 0 20px;font-size:13.5px}th{background:var(--navy);color:white;text-align:left;padding:9px 10px}td{border-bottom:1px solid var(--line);padding:9px 10px;vertical-align:top}.num{text-align:right;white-space:nowrap}.bad{color:var(--red);font-weight:800}.good{color:var(--green);font-weight:800}.warning{color:var(--warn);font-weight:800}ul{padding-left:20px}footer{margin-top:44px;padding-top:22px;border-top:1px solid var(--line);color:var(--muted);font-size:12px}@media(max-width:760px){.page{margin:0}header,main{padding-left:22px;padding-right:22px}header h1{font-size:34px}.grid{grid-template-columns:1fr}table{display:block;overflow-x:auto}}
"""


def _money(v, decimals=2):
    return f"${v:,.{decimals}f}"


def build_memo(path: Path | None = None) -> Path:
    d = build_data()
    sc = d["scenarios"]
    base_eps = next(r for r in sc["rows"] if r["key"] == "base")["eps"]
    adv_usd = INPUTS["adv_shares_30d"] * INPUTS["spot"]
    scenario_rows = "".join(
        f"<tr><td><b>{r['label']}</b><br><span class='sub'>{r['basis']}</span></td>"
        f"<td class='num'>{r['case_da']/1e9:.1f}</td><td class='num'>{r['ebitda_offset']:.0%}</td>"
        f"<td class='num'>{_money(r['eps'])}</td><td class='num'>{r['multiple']:.1f}x</td>"
        f"<td class='num'>{_money(r['target'])}</td><td class='num'>{r['return']:+.1%}</td>"
        f"<td class='num'>{r['probability']:.0%}</td></tr>" for r in sc["rows"])
    gate_rows = "".join(f"<tr><td><b>{g}</b></td><td class='bad'>{s}</td><td>{m}</td></tr>" for g, s, m in IMPLEMENTATION_GATES)
    dep_rows = "".join(
        f"<tr><td>{y}</td><td class='num'>{INPUTS['capex'][i]/1e9:.1f}</td>"
        f"<td class='num'>{d['depreciation']['annual_legacy'][y]/1e9:.1f}</td>"
        f"<td class='num'>{d['depreciation']['annual_forecast'][y]/1e9:.1f}</td>"
        f"<td class='num'><b>{d['depreciation']['annual'][y]/1e9:.1f}</b></td></tr>"
        for i, y in enumerate(YEARS))
    rule_rows = "".join(f"<tr><td><b>{a}</b></td><td>{b}</td></tr>" for a, b in ACTION_RULES)
    source_rows = "".join(
        f"<tr><td>{s[0]}</td><td>{s[1]}</td><td>{s[2]}</td><td>{s[3]}</td>"
        f"<td><a href='{s[6]}'>{s[4]}</a></td><td>{s[5]}</td></tr>" if s[6] else
        f"<tr><td>{s[0]}</td><td>{s[1]}</td><td>{s[2]}</td><td>{s[3]}</td><td>{s[4]}</td><td>{s[5]}</td></tr>"
        for s in SOURCES)
    html = f"""<!doctype html><html lang='en'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><link rel='icon' href='data:,'><title>Alphabet (GOOGL) — PM readiness memo</title><style>{CSS}</style></head><body><div class='page'>
<header><div class='eyebrow'>Alphabet · GOOGL · PM readiness memo</div><h1>Wait for proof. No position.</h1>
<p class='dek'>The D&A/EPS hypothesis is monitorable, but it is not yet an implementable short. Initiation requires a reported D&A surprise that survives the EBITDA offset, produces observable FY2027 estimate cuts, retains at least 20% net downside, and clears every borrow, crowding, options and hedge gate.</p>
<div class='chips'><span class='chip'>Spot {_money(INPUTS['spot'])}</span><span class='chip'>Current size 0.0%</span><span class='chip'>18-month research horizon</span><span class='chip'>As of 7 Aug 2026</span></div></header><main>
<div class='callout stop'><b>PM decision:</b> circulate as a research-monitoring memo, not a trade ticket. Calculation integrity passes the rebuilt controls; trade implementation does not.</div>
<div class='grid'><div class='card'><div class='sub'>Primary scenario EV</div><div class='metric'>{_money(sc['expected_value'])}</div><div class='sub'>{sc['expected_return']:+.1%} vs spot; illustrative only</div></div>
<div class='card'><div class='sub'>DCF cross-check</div><div class='metric'>{_money(d['dcf']['value_per_share'])}</div><div class='sub'>{d['dcf']['tv_pct_ev']:.0%} terminal value; risk frame</div></div>
<div class='card'><div class='sub'>FY2027 EPS thought experiment</div><div class='metric'>{_money(base_eps)}</div><div class='sub'>vs frozen FMP {_money(INPUTS['street_fy27_eps'])}; source conflict open</div></div></div>

<h2>1. One causal thesis</h2><p><b>Alphabet's AI capex converts into GAAP D&A faster than consensus EPS incorporates it; if EBITDA growth does not absorb the charge, quarterly prints force FY2027-28 EPS cuts and a multiple reset.</b></p>
<p>D&A is an earnings-revision catalyst, not the direct DCF downside. In FCFF it is added back; the cash leaves through capex. A rising D&A line is therefore not confirmation by itself. The signal is the portion that remains unoffset in EBIT, reaches EPS, changes the analyst distribution, and then changes the multiple.</p>

<h2>2. Implementation gate</h2><table><tr><th>Gate</th><th>Status</th><th>What remains</th></tr>{gate_rows}</table>
<p class='sub'>Robinhood read-only market data shows 30-day GOOGL ADV of {INPUTS['adv_shares_30d']/1e6:.1f}m shares, approximately ${adv_usd/1e9:.1f}bn at spot. That supports ordinary underlying liquidity; it does not establish short locate, carry, squeeze capacity, option execution, hedge basis, or an exit plan for a specific book.</p>

<h2>3. What is priced in—and what is not proven</h2><p>At spot, the frozen FMP FY2027 EPS feed implies {INPUTS['spot']/INPUTS['street_fy27_eps']:.1f}x. The base EPS thought experiment implies {INPUTS['spot']/base_eps:.1f}x. Those are observable arithmetic, not proof that the market explicitly believes any single WACC, terminal margin, or revenue CAGR.</p>
<div class='callout warn'><b>Consensus control failure remains open.</b> FMP carries FY2027 EPS of {_money(INPUTS['street_fy27_eps'])} from {INPUTS['street_fy27_n']} estimates with a {_money(INPUTS['street_fy27_eps_low'])}-{_money(INPUTS['street_fy27_eps_high'])} range. A separately cited $14.20-$14.68 range does not bracket $15.01. Until provider, timestamp, mean/median, estimate dates and revision history reconcile, the annual EPS bridge is a research scenario—not a variant perception we would risk capital against.</div>

<h2>4. Placed-in-service D&A</h2><p>The rebuilt schedule starts from reported Q2 D&A of ${INPUTS['q2_da']/1e9:.1f}bn and adds depreciation only after explicit commissioning lags. Asset mix, lives and lags are blue analyst assumptions in the workbook. Historical/opening D&A is separated from forecast-vintage D&A; the memo no longer claims all modeled depreciation comes from forecast capex.</p>
<table><tr><th>Year</th><th class='num'>Capex</th><th class='num'>Opening-vintage D&A</th><th class='num'>Forecast-vintage D&A</th><th class='num'>Total D&A</th></tr>{dep_rows}</table>
<p>FY2026 capex is $200bn, the midpoint of the $195-205bn guide. FY2027 is an explicit $230bn assumption, +15%, consistent with management's direction that spending would increase significantly. There is no kill criterion that the base case trips on day one.</p>

<h2>5. Illustrative scenario skew</h2><table><tr><th>Case</th><th class='num'>FY27 D&A</th><th class='num'>EBITDA offset</th><th class='num'>FY27 EPS</th><th class='num'>Multiple</th><th class='num'>Target</th><th class='num'>Return</th><th class='num'>Probability</th></tr>{scenario_rows}</table>
<p>Scenario values are EPS × multiple and reflect the actual causal chain. The DCF is a cross-check, not the target-setting mechanism. Expected value is shown because it is formula-linked, but it does not clear the trade gate and does not produce a position size.</p>

<h2>6. Quarterly catalyst bridge</h2><table><tr><th>Print</th><th>Event status</th><th class='num'>Model D&A</th><th>Frozen consensus D&A</th><th>Required result</th></tr>
<tr><td>Q3 2026</td><td>Tentative / date not company-confirmed here</td><td class='num'>${d['depreciation']['total'][0]/1e9:.1f}bn</td><td class='bad'>Missing</td><td>D&A surprise, EBITDA offset, FY2027 revision at 1/5/10 days</td></tr>
<tr><td>Q4 2026</td><td>Tentative</td><td class='num'>${d['depreciation']['total'][1]/1e9:.1f}bn</td><td class='bad'>Missing</td><td>Second print; useful-life and FY2027 capex disclosures</td></tr>
<tr><td>Q1 2027</td><td>Tentative</td><td class='num'>${d['depreciation']['total'][2]/1e9:.1f}bn</td><td class='bad'>Missing</td><td>Catalyst expiry test if consensus EPS is flat/up</td></tr></table>
<p class='sub'>Before each print freeze quarterly EBITDA, EBIT, implied D&A, tax, below-line income, diluted shares and EPS; record actual D&A, EBITDA surprise/offset, FY2027 EPS before/after, stock reaction and post-print multiple.</p>

<h2>7. Conditional action rules</h2><table><tr><th>Action</th><th>Rule</th></tr>{rule_rows}</table>

<h2>8. Valuation and capital-structure controls</h2><ul>
<li><b>Dated stub:</b> the June 30 bridge is rolled to August 7 with an explicitly estimated interim FCFF; only August 8 onward is discounted. Reported H1 is not counted again.</li>
<li><b>Terminal normalization:</b> next-period NOPAT is reduced by g / terminal ROIC reinvestment before applying Gordon growth. Base WACC is {d['wacc']['wacc']:.2%}, built from a 4.65% Treasury rate, 0.95 beta, 4.50% ERP and market-value capital weights.</li>
<li><b>Shares:</b> {d['share_bridge']['total']/1e9:.3f}bn point diluted shares = 12.230bn point-in-time common + ordinary dilution + minimum mandatory-convertible dilution. Preferred value is not deducted again.</li>
<li><b>Financing language:</b> the $40bn ATM is registered capacity available but unsold at June 30, not charter-authorized capital. Unsupported aggregate financing claims have been removed; completed common and preferred issuance is separated from preliminary or undrawn capacity.</li>
<li><b>Multiple:</b> enterprise earnings use EV/NOPAT; there is no price/NOPAT-per-share "economic P/E."</li></ul>

<h2>Appendix A. Source ledger</h2><table><tr><th>ID</th><th>Type</th><th>Metric</th><th>As of</th><th>Document</th><th>Location</th></tr>{source_rows}</table>
<h2>Appendix B. Known limitations</h2><ul><li>No broker-level quarterly consensus ledger or revision history.</li><li>No live borrow/carry, short-interest/crowding, option-chain execution or hedge analysis.</li><li>Capex asset mix, useful lives, commissioning lags, beta, ERP, terminal ROIC, scenario probabilities and exit multiples are analyst assumptions.</li><li>The opening-vintage D&A run-rate is held flat through the explicit horizon because remaining lives by historical vintage are not disclosed.</li></ul>
<footer>Prepared 8 August 2026 from public information. Market data as of the 7 August 2026 session; financial cutoff 30 June 2026. Analytical research only—not investment advice, a recommendation, or a solicitation.</footer>
</main></div></body></html>"""
    target = path or ROOT / "alphabet" / "memo.html"
    target.write_text(html)
    return target


def main():
    data = build_data()
    print(f"Alphabet PM model: WACC {data['wacc']['wacc']:.2%}; DCF ${data['dcf']['value_per_share']:.2f}; scenario EV ${data['scenarios']['expected_value']:.2f}")
    print("wrote", build_workbook())
    print("wrote", build_memo())


if __name__ == "__main__":
    main()
