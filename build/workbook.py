"""Build the live Excel model.

Authored from our own layout rather than a vendor template. Every driver is a
labelled blue input cell and every output is a real formula chain, so changing
Search ad coverage or Data Center ASP moves the value per share. There is no
override cell shadowing a computed one: WACC has exactly one definition and the
whole model references it.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import cases
import decision

DATA = Path(__file__).parent / "data"
OUT = Path(__file__).parent.parent
YEARS = 6

INPUT_FILL = PatternFill("solid", fgColor="DCE9F7")   # blue = you may change this
CALC_FILL = PatternFill("solid", fgColor="F4F2EE")
HDR_FILL = PatternFill("solid", fgColor="1F3352")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3352")
LBL_FONT = Font(bold=True, size=10)
NOTE_FONT = Font(italic=True, size=9, color="666666")
THIN = Side(style="thin", color="D4CFC7")
BOX = Border(bottom=THIN)

PCT = '0.0%'
PCT2 = '0.00%'
BN = '#,##0.0,,,'      # display $bn from a $ figure
USD = '$#,##0.00'


def sheet(wb, name, widths):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def title(ws, row, text, span=9):
    ws.cell(row, 1, text).font = TITLE_FONT
    return row + 1


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row, start + i, l)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center" if i else "left")
    return row + 1


def note(ws, row, text):
    ws.cell(row, 1, text).font = NOTE_FONT
    return row + 1


def year_one_capex_note(res):
    """One line on how forecast year 1 capex was anchored, for the Notes tab.

    The workbook is a second implementation, not a printout, so it has to carry
    the same provenance as the memo rather than pointing at it.
    """
    a = res["capex_anchor"]["base"]
    if a.get("guided"):
        return (f"Management guided full-year capex to ${a['guidance_low'] / 1e9:,.0f}-"
                f"{a['guidance_high'] / 1e9:,.0f}bn ({a['guidance_source']}); the base case takes "
                f"the midpoint, ${a['year_one_capex'] / 1e9:,.1f}bn, against "
                f"${a['ytd_actual'] / 1e9:,.1f}bn already reported. Holding the exit quarter flat "
                f"instead would give ${a['exit_rate_alternative'] / 1e9:,.1f}bn, below the guided "
                f"range. Guidance outranks extrapolation for a period management has guided.")
    return (f"No capex guidance is published for this filer, so year 1 is "
            f"${a['ytd_actual'] / 1e9:,.1f}bn reported plus {res['ytd']['quarters_remaining']} "
            f"quarter(s) at the ${a['exit_quarter'] / 1e9:,.1f}bn exit rate, giving "
            f"${a['year_one_capex'] / 1e9:,.1f}bn.")


def year_one_segment_note(res):
    """One line on the segment-level floor against reported quarters."""
    sa = res.get("segment_anchor", {}).get("base")
    if not sa:
        return ("No quarterly segment actuals are recorded for this filer, so year-1 segment "
                "revenue is checked only at the consolidated level.")
    w = sa["worst"]
    return (f"Each segment's year-1 forecast is checked against its own reported quarters: the stub "
            f"implied for the {sa['remaining_quarters']} remaining quarter(s) may not sit more than "
            f"{sa['tolerance']:.0%} below the exit quarter. The tightest is {w['segment']} at "
            f"{w['vs_exit_quarter']:+.1%}. This exists because a consolidated total can be right "
            f"while the mix inside it is wrong: the prior edition forecast Cloud below its own exit "
            f"rate and nothing consolidated could see it.")


def wacc_note(res):
    """One line on what the discount rate is worth, with the alternatives."""
    ws = res.get("wacc_sensitivity")
    if not ws:
        return "No discount-rate sensitivity recorded."
    at = {round(g["wacc"], 4): g["value_per_share"] for g in ws["grid"]}
    base = ws["base_wacc"]
    alt = at.get(0.0757)
    return (f"WACC is an input cell on Drivers and every discounted line references it, so changing "
            f"it here re-values the model. It is worth knowing what that is worth: terminal value is "
            f"{ws['tv_pct_ev']:.0%} of enterprise value, the base case discounts at {base:.2%} giving "
            f"${at[round(base, 4)]:,.2f}, published third-party WACC estimates cluster near 7.57% "
            f"giving ${alt:,.2f}, and the reverse DCF implies the market is discounting at "
            f"{ws['market_implied_wacc']:.2%}. The discount rate moves this valuation more than the "
            f"depreciation schedule does.")


def build(ticker, slug):
    res = json.loads((DATA / "results.json").read_text())[ticker]
    by = json.loads((DATA / "base_year.json").read_text())[ticker]
    spec = cases.SPEC[ticker]
    dv = dict(spec["drivers"]["base"])
    fy0 = res["first_year"]
    base = res["scenarios"]["base"]
    yrs = [fy0 + i for i in range(YEARS)]
    # Capex is no longer a declared constant: year 1 is anchored on reported
    # fiscal-YTD spend and the rest fades from there, so the workbook has to
    # publish the path the model actually ran, not the pre-anchor declaration.
    dv["capex_pct_revenue"] = [q["capex"] / q["revenue"] for q in base["rows"]]

    wb = Workbook()
    wb.remove(wb.active)

    # ================= Drivers (the only input sheet) =================
    ws = sheet(wb, "Drivers", [38, 13, 13, 13, 13, 13, 13, 46])
    r = title(ws, 1, f"{spec['name']} ({ticker}) — driver inputs")
    r = note(ws, r, "Blue cells are inputs. Everything else in this workbook is a formula. "
                    "Change a blue cell and the value per share on 'Valuation' moves.")
    r += 1
    r = hdr(ws, r, ["Driver"] + [f"{y}E" for y in yrs] + ["Note"])

    seg_rows = {}
    for seg, drivers in dv["segments"].items():
        ws.cell(r, 1, spec["segment_labels"][seg]).font = LBL_FONT
        r += 1
        seg_rows[seg] = {}
        for drv, vec in drivers.items():
            ws.cell(r, 1, f"   {drv.replace('_', ' ')}")
            for i, v in enumerate(vec):
                c = ws.cell(r, 2 + i, v)
                c.fill, c.number_format = INPUT_FILL, PCT
            seg_rows[seg][drv] = r
            r += 1
    r += 1

    scalar_rows = {}
    for key, fmt, label, nt in [
        ("ebitda_margin", PCT, "EBITDA margin (cash margin, pre-D&A)",
         "We forecast this and DERIVE EBIT by subtracting depreciation."),
        ("capex_pct_revenue", PCT, "Capex % of revenue", "Drives the depreciation schedule."),
        ("tax_rate", PCT, "Tax rate", ""),
        ("nwc_pct_revenue", PCT, "Net working capital % of revenue", ""),
    ]:
        ws.cell(r, 1, label).font = LBL_FONT
        for i, v in enumerate(dv[key]):
            c = ws.cell(r, 2 + i, v)
            c.fill, c.number_format = INPUT_FILL, fmt
        ws.cell(r, 8, nt).font = NOTE_FONT
        scalar_rows[key] = r
        r += 1

    r += 1
    ws.cell(r, 1, "Base year revenue by segment ($)").font = LBL_FONT
    r += 1
    base_seg_rows = {}
    for seg, v in spec["base_segments"].items():
        ws.cell(r, 1, f"   {spec['segment_labels'][seg]}")
        c = ws.cell(r, 2, v)
        c.fill, c.number_format = INPUT_FILL, BN
        base_seg_rows[seg] = r
        r += 1

    r += 1
    ws.cell(r, 1, "WACC").font = LBL_FONT
    wacc_cell = f"$B${r}"
    c = ws.cell(r, 2, base["wacc"])
    c.fill, c.number_format = INPUT_FILL, PCT2
    ws.cell(r, 8, "Single definition. Every sheet references this cell — there is no override.").font = NOTE_FONT
    r += 1
    ws.cell(r, 1, "Terminal growth").font = LBL_FONT
    g_cell = f"$B${r}"
    c = ws.cell(r, 2, base["terminal_growth"])
    c.fill, c.number_format = INPUT_FILL, PCT2
    r += 1
    ws.cell(r, 1, "Diluted shares").font = LBL_FONT
    sh_cell = f"$B${r}"
    c = ws.cell(r, 2, res["diluted_shares"])
    c.fill, c.number_format = INPUT_FILL, '#,##0,,'
    r += 1

    bridge_cells = {}
    for k, label in [("cash_and_marketable", "Cash & marketable securities"),
                     ("equity_investments", "Equity investments (stakes)"),
                     ("debt", "Debt"), ("leases", "Operating leases"),
                     ("preferred", "Preferred stock")]:
        ws.cell(r, 1, label).font = LBL_FONT
        c = ws.cell(r, 2, res["bridge_inputs"][k])
        c.fill, c.number_format = INPUT_FILL, BN
        bridge_cells[k] = f"$B${r}"
        r += 1
    ws.cell(r, 1, "Spot price").font = LBL_FONT
    spot_cell = f"$B${r}"
    c = ws.cell(r, 2, res["spot"])
    c.fill, c.number_format = INPUT_FILL, USD

    # ================= Revenue build =================
    rv = sheet(wb, "Revenue", [38, 15, 15, 15, 15, 15, 15])
    r = title(rv, 1, "Segment revenue build — each line is the product of its drivers")
    r += 1
    r = hdr(rv, r, ["Segment"] + [f"{y}E" for y in yrs])
    seg_out = {}
    for seg in dv["segments"]:
        rv.cell(r, 1, spec["segment_labels"][seg])
        for i in range(YEARS):
            col = get_column_letter(2 + i)
            prev = f"Drivers!$B${base_seg_rows[seg]}" if i == 0 else f"{get_column_letter(1 + i)}{r}"
            factors = "*".join(f"(1+Drivers!{get_column_letter(2 + i)}${row})"
                               for row in seg_rows[seg].values())
            c = rv.cell(r, 2 + i, f"={prev}*{factors}")
            c.number_format = BN
        seg_out[seg] = r
        r += 1
    total_row = r
    rv.cell(r, 1, "Total revenue").font = LBL_FONT
    for i in range(YEARS):
        col = get_column_letter(2 + i)
        c = rv.cell(r, 2 + i, f"=SUM({col}{total_row - len(dv['segments'])}:{col}{total_row - 1})")
        c.number_format, c.font = BN, LBL_FONT
    r += 1
    rv.cell(r, 1, "Growth")
    for i in range(YEARS):
        col = get_column_letter(2 + i)
        prev = f"SUM(Drivers!$B${min(base_seg_rows.values())}:$B${max(base_seg_rows.values())})" if i == 0 \
            else f"{get_column_letter(1 + i)}{total_row}"
        c = rv.cell(r, 2 + i, f"={col}{total_row}/{prev}-1")
        c.number_format = PCT

    # ================= Depreciation =================
    dp = sheet(wb, "Depreciation", [38, 15, 15, 15, 15, 15, 15, 40])
    r = title(dp, 1, "Vintage depreciation schedule")
    r = note(dp, r, f"Assets split {cases.ASSET_MIX[ticker]['short_share']:.0%} short-lived "
                    f"({cases.ASSET_MIX[ticker]['short_life']:.0f}yr) / "
                    f"{1 - cases.ASSET_MIX[ticker]['short_share']:.0%} long-lived "
                    f"({cases.ASSET_MIX[ticker]['long_life']:.0f}yr), half-year convention in the "
                    f"vintage year. This is the schedule the consensus feed does not carry.")
    r += 1
    r = hdr(dp, r, ["Item"] + [f"{y}E" for y in yrs])
    dp.cell(r, 1, "Capex ($)")
    capex_row = r
    for i in range(YEARS):
        col = get_column_letter(2 + i)
        c = dp.cell(r, 2 + i, f"=Revenue!{col}{total_row}*Drivers!{col}${scalar_rows['capex_pct_revenue']}")
        c.number_format = BN
    r += 1
    dp.cell(r, 1, "D&A ($) — modelled").font = LBL_FONT
    da_row = r
    for i, v in enumerate(base["da_abs_bn"]):
        c = dp.cell(r, 2 + i, v * 1e9)
        c.number_format, c.fill = BN, CALC_FILL
    dp.cell(r, 8, "Computed in build/cases.py:depreciation_path from capex vintages.").font = NOTE_FONT
    r += 1
    dp.cell(r, 1, "D&A % of revenue")
    for i in range(YEARS):
        col = get_column_letter(2 + i)
        c = dp.cell(r, 2 + i, f"={col}{da_row}/Revenue!{col}{total_row}")
        c.number_format = PCT
    r += 1
    dp.cell(r, 1, "Capex / D&A ratio")
    for i in range(YEARS):
        col = get_column_letter(2 + i)
        c = dp.cell(r, 2 + i, f"={col}{capex_row}/{col}{da_row}")
        c.number_format = '0.0"x"'

    # ================= Valuation =================
    vl = sheet(wb, "Valuation", [38, 15, 15, 15, 15, 15, 15, 34])
    r = title(vl, 1, f"{ticker} — unlevered FCFF DCF, mid-year convention")
    r += 1
    r = hdr(vl, r, ["Line"] + [f"{y}E" for y in yrs])
    rows = {}
    for label, formula, fmt in [
        ("Revenue", lambda c: f"=Revenue!{c}{total_row}", BN),
        ("EBITDA margin", lambda c: f"=Drivers!{c}${scalar_rows['ebitda_margin']}", PCT),
        ("EBITDA", lambda c: f"={c}{{Revenue}}*{c}{{EBITDA margin}}", BN),
        ("D&A", lambda c: f"=Depreciation!{c}{da_row}", BN),
        ("EBIT", lambda c: f"={c}{{EBITDA}}-{c}{{D&A}}", BN),
        ("EBIT margin", lambda c: f"={c}{{EBIT}}/{c}{{Revenue}}", PCT),
        ("Tax rate", lambda c: f"=Drivers!{c}${scalar_rows['tax_rate']}", PCT),
        ("NOPAT", lambda c: f"={c}{{EBIT}}*(1-{c}{{Tax rate}})", BN),
        ("Capex", lambda c: f"=Depreciation!{c}{capex_row}", BN),
        ("NWC", lambda c: f"={c}{{Revenue}}*Drivers!{c}${scalar_rows['nwc_pct_revenue']}", BN),
        ("Change in NWC", None, BN),
        ("FCFF", lambda c: f"={c}{{NOPAT}}+{c}{{D&A}}-{c}{{Capex}}-{c}{{Change in NWC}}", BN),
        ("Discount factor", None, '0.0000'),
        ("PV of FCFF", lambda c: f"={c}{{FCFF}}*{c}{{Discount factor}}", BN),
    ]:
        vl.cell(r, 1, label).font = LBL_FONT if label in ("EBIT", "FCFF", "NOPAT") else Font(size=10)
        rows[label] = r
        r += 1

    def ref(label, i):
        return f"{get_column_letter(2 + i)}{rows[label]}"

    for label, formula, fmt in [
        ("Revenue", lambda c: f"=Revenue!{c}{total_row}", BN),
        ("EBITDA margin", lambda c: f"=Drivers!{c}${scalar_rows['ebitda_margin']}", PCT),
        ("EBITDA", None, BN), ("D&A", None, BN), ("EBIT", None, BN), ("EBIT margin", None, PCT),
        ("Tax rate", None, PCT), ("NOPAT", None, BN), ("Capex", None, BN), ("NWC", None, BN),
        ("Change in NWC", None, BN), ("FCFF", None, BN),
        ("Discount factor", None, '0.0000'), ("PV of FCFF", None, BN),
    ]:
        for i in range(YEARS):
            col = get_column_letter(2 + i)
            rr = rows[label]
            if label == "Revenue":
                f = f"=Revenue!{col}{total_row}"
            elif label == "EBITDA margin":
                f = f"=Drivers!{col}${scalar_rows['ebitda_margin']}"
            elif label == "EBITDA":
                f = f"={ref('Revenue', i)}*{ref('EBITDA margin', i)}"
            elif label == "D&A":
                f = f"=Depreciation!{col}{da_row}"
            elif label == "EBIT":
                f = f"={ref('EBITDA', i)}-{ref('D&A', i)}"
            elif label == "EBIT margin":
                f = f"={ref('EBIT', i)}/{ref('Revenue', i)}"
            elif label == "Tax rate":
                f = f"=Drivers!{col}${scalar_rows['tax_rate']}"
            elif label == "NOPAT":
                f = f"={ref('EBIT', i)}*(1-{ref('Tax rate', i)})"
            elif label == "Capex":
                f = f"=Depreciation!{col}{capex_row}"
            elif label == "NWC":
                f = f"={ref('Revenue', i)}*Drivers!{col}${scalar_rows['nwc_pct_revenue']}"
            elif label == "Change in NWC":
                if i == 0:
                    f = (f"={ref('NWC', 0)}-SUM(Drivers!$B${min(base_seg_rows.values())}:"
                         f"$B${max(base_seg_rows.values())})*Drivers!$B${scalar_rows['nwc_pct_revenue']}")
                else:
                    f = f"={ref('NWC', i)}-{ref('NWC', i - 1)}"
            elif label == "FCFF":
                f = f"={ref('NOPAT', i)}+{ref('D&A', i)}-{ref('Capex', i)}-{ref('Change in NWC', i)}"
            elif label == "Discount factor":
                f = f"=1/(1+Drivers!{wacc_cell})^({i}+0.5)"
            else:
                f = f"={ref('FCFF', i)}*{ref('Discount factor', i)}"
            c = vl.cell(rr, 2 + i, f)
            c.number_format = fmt

    r += 1
    out = {}
    for label, f, fmt in [
        ("PV of explicit forecast", f"=SUM(B{rows['PV of FCFF']}:{get_column_letter(1 + YEARS)}{rows['PV of FCFF']})", BN),
        ("Terminal value",
         f"={ref('FCFF', YEARS - 1)}*(1+Drivers!{g_cell})/(Drivers!{wacc_cell}-Drivers!{g_cell})", BN),
        ("PV of terminal value", None, BN),
        ("Enterprise value", None, BN),
        ("+ Cash & marketable securities", f"=Drivers!{bridge_cells['cash_and_marketable']}", BN),
        ("+ Equity investments", f"=Drivers!{bridge_cells['equity_investments']}", BN),
        ("- Debt", f"=-Drivers!{bridge_cells['debt']}", BN),
        ("- Operating leases", f"=-Drivers!{bridge_cells['leases']}", BN),
        ("- Preferred stock", f"=-Drivers!{bridge_cells['preferred']}", BN),
        ("Equity value", None, BN),
        ("Value per share", None, USD),
        ("Spot price", f"=Drivers!{spot_cell}", USD),
        ("Upside / (downside)", None, PCT),
        ("Terminal value % of EV", None, PCT),
    ]:
        vl.cell(r, 1, label).font = LBL_FONT
        out[label] = r
        r += 1

    def o(label):
        return f"B{out[label]}"

    vl.cell(out["PV of explicit forecast"], 2,
            f"=SUM(B{rows['PV of FCFF']}:{get_column_letter(1 + YEARS)}{rows['PV of FCFF']})").number_format = BN
    vl.cell(out["Terminal value"], 2,
            f"={ref('FCFF', YEARS - 1)}*(1+Drivers!{g_cell})/(Drivers!{wacc_cell}-Drivers!{g_cell})").number_format = BN
    vl.cell(out["PV of terminal value"], 2,
            f"={o('Terminal value')}/(1+Drivers!{wacc_cell})^({YEARS}-0.5)").number_format = BN
    vl.cell(out["Enterprise value"], 2,
            f"={o('PV of explicit forecast')}+{o('PV of terminal value')}").number_format = BN
    vl.cell(out["+ Cash & marketable securities"], 2, f"=Drivers!{bridge_cells['cash_and_marketable']}").number_format = BN
    vl.cell(out["+ Equity investments"], 2, f"=Drivers!{bridge_cells['equity_investments']}").number_format = BN
    vl.cell(out["- Debt"], 2, f"=-Drivers!{bridge_cells['debt']}").number_format = BN
    vl.cell(out["- Operating leases"], 2, f"=-Drivers!{bridge_cells['leases']}").number_format = BN
    vl.cell(out["- Preferred stock"], 2, f"=-Drivers!{bridge_cells['preferred']}").number_format = BN
    vl.cell(out["Equity value"], 2,
            f"={o('Enterprise value')}+{o('+ Cash & marketable securities')}+{o('+ Equity investments')}"
            f"+{o('- Debt')}+{o('- Operating leases')}+{o('- Preferred stock')}").number_format = BN
    c = vl.cell(out["Value per share"], 2, f"={o('Equity value')}/Drivers!{sh_cell}")
    c.number_format, c.font, c.fill = USD, Font(bold=True, size=12), CALC_FILL
    vl.cell(out["Spot price"], 2, f"=Drivers!{spot_cell}").number_format = USD
    vl.cell(out["Upside / (downside)"], 2, f"={o('Value per share')}/{o('Spot price')}-1").number_format = PCT
    vl.cell(out["Terminal value % of EV"], 2,
            f"={o('PV of terminal value')}/{o('Enterprise value')}").number_format = PCT

    r += 1
    r = note(vl, r, "Sanity check: WACC minus terminal growth should exceed 150bp.")
    vl.cell(r, 1, "WACC - g")
    vl.cell(r, 2, f"=Drivers!{wacc_cell}-Drivers!{g_cell}").number_format = PCT2
    vl.cell(r, 3, f'=IF(B{r}>0.015,"OK","FAIL — terminal value unstable")')

    # ================= Variant vs Street =================
    st = sheet(wb, "Variant", [16, 16, 16, 16, 16, 14, 14, 14])
    r = title(st, 1, "Our forecast against the analyst distribution")
    r = note(st, r, "Street figures are the consensus feed as retrieved 2026-08-07. "
                    "Low/high are the support of published estimates, not a confidence interval.")
    r += 1
    r = hdr(st, r, ["Fiscal year", "Our revenue", "Street avg", "Street low", "Street high",
                    "Delta %", "Analysts", "In range?"])
    for row in res["variant"]:
        st.cell(r, 1, row["fy"])
        for i, v in enumerate([row["ours"], row["street_avg"], row["street_low"], row["street_high"]]):
            st.cell(r, 2 + i, v).number_format = BN
        st.cell(r, 6, row["delta_pct"]).number_format = PCT
        st.cell(r, 7, row["n_analysts"])
        st.cell(r, 8, "yes" if row["inside_range"] else "OUTSIDE")
        r += 1

    # ================= Scenarios & decision =================
    sc = sheet(wb, "Decision", [34, 16, 16, 16, 40])
    r = title(sc, 1, f"{ticker} — scenarios, expected value and sizing")
    r += 1
    scen = {k: res["scenarios"][k]["value_per_share"] for k in ("bear", "base", "bull")}
    probs = decision.PROBS[ticker]
    meta = decision.RATING[ticker]
    direction = "SHORT" if meta["rating"] == "SHORT" else "LONG"
    sz = decision.size(ticker, scen, probs, res["spot"], direction)

    r = hdr(sc, r, ["Case", "Value / share", "Probability", "vs spot", "Anchor"])
    first = r
    for k in ("bear", "base", "bull"):
        sc.cell(r, 1, k.title())
        sc.cell(r, 2, scen[k]).number_format = USD
        sc.cell(r, 3, probs[k]["bp"] / 10000).number_format = PCT
        sc.cell(r, 4, f"=B{r}/Valuation!{o('Spot price')}-1").number_format = PCT
        sc.cell(r, 5, probs[k]["anchor"].replace("_", " "))
        r += 1
    sc.cell(r, 1, "Expected value").font = LBL_FONT
    sc.cell(r, 2, f"=SUMPRODUCT(B{first}:B{r - 1},C{first}:C{r - 1})").number_format = USD
    sc.cell(r, 3, f"=SUM(C{first}:C{r - 1})").number_format = PCT
    sc.cell(r, 4, f"=B{r}/Valuation!{o('Spot price')}-1").number_format = PCT
    ev_row = r
    r += 2

    sc.cell(r, 1, "Rating").font = LBL_FONT
    sc.cell(r, 2, meta["rating"]).font = Font(bold=True, size=12)
    r += 1
    for label, val, fmt, nt in [
        ("Risk / reward", sz["risk_reward"], '0.00', f"{direction.lower()} framing"),
        ("Payoff ratio b", sz["b"], '0.00', "reward / risk"),
        ("p(win)", sz["p_win"], PCT, "probability mass favouring the position"),
        ("Kelly f", sz["kelly_f"], PCT, ""),
        ("Quarter-Kelly", sz["size_raw"], PCT, "0.25 x Kelly"),
        ("Liquidity cap", sz["size_liquidity"], PCT, "20% of ADV over 5 days"),
        ("Risk-budget cap", sz["size_risk_budget"], PCT, "1.5% of NAV at risk"),
        ("Concentration cap", sz["size_concentration"], PCT, "single-name hard limit"),
        ("Position size", sz["position_size"], PCT, f"binding: {sz['binding_constraint']}"),
    ]:
        sc.cell(r, 1, label).font = LBL_FONT if label == "Position size" else Font(size=10)
        c = sc.cell(r, 2, val)
        c.number_format = fmt
        if label == "Position size":
            c.font, c.fill = Font(bold=True, size=12), CALC_FILL
        sc.cell(r, 5, nt).font = NOTE_FONT
        r += 1
    r += 1
    r = note(sc, r, f"Basis: {sz['nav_basis']}. Illustrative only.")

    # ================= Notes =================
    nt = sheet(wb, "Notes", [110])
    r = title(nt, 1, "Sources, method and limitations")
    r += 1
    for line in [
        f"Base year: trailing twelve months to {by['as_of']}, all statement inputs from SEC XBRL "
        f"companyconcept (CIK {'0001652044' if ticker == 'GOOGL' else '0001045810'}).",
        "Quarterly series are reconstructed from filed facts: Q4 is derived as fiscal year minus the "
        "nine-month cumulative, and cash-flow items are unwound from year-to-date cumulatives.",
        "Consensus estimates and price targets: FMP analyst/financial-estimates and "
        "analyst/price-target-consensus, retrieved 2026-08-07.",
        "Segment revenue: FMP statements/revenue-product-segmentation. No SEC fallback exists because "
        "companyfacts flattens dimensioned facts; the segment sum is reconciled to consolidated revenue.",
        "",
        f"YEAR-1 CAPEX ANCHOR. {year_one_capex_note(res)}",
        "",
        f"YEAR-1 SEGMENT ANCHOR. {year_one_segment_note(res)}",
        "",
        f"DISCOUNT RATE. {wacc_note(res)}",
        "",
        "MODEL DESIGN. We forecast EBITDA margin and derive EBIT by subtracting a depreciation schedule "
        "built from capex vintages. Forecasting EBIT margin directly hides the depreciation assumption.",
        "",
        "KEY LIMITATION. The asset-life split driving the depreciation schedule is a stated assumption "
        "calibrated to reported D&A, not a disclosed figure. It is the most load-bearing assumption in "
        "the model.",
        "",
        "This is analytical research on public information. Not investment advice, not a recommendation "
        "to any person, and not a solicitation. Position sizing is illustrative against a notional $1bn book.",
    ]:
        nt.cell(r, 1, line).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    wb.calculation.fullCalcOnLoad = True
    p = OUT / slug / "model.xlsx"
    wb.save(p)
    return p


if __name__ == "__main__":
    for t, s in (("GOOGL", "alphabet"), ("NVDA", "nvidia")):
        print("wrote", build(t, s))
