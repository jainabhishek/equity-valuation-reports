"""Independent verification: recalculate the workbook and compare it to the Python model.

This is the check that matters. The workbook and the Python engine are two
separate implementations of the same model; if they disagree, one of them is
wrong and the published number cannot be trusted. Comparing Python to itself
would prove nothing.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import formulas
from openpyxl import load_workbook

import alphabet_pm

DATA = Path(__file__).parent / "data"
OUT = Path(__file__).parent.parent
TOL = 0.01  # cents per share


def solve(path):
    xl = formulas.ExcelModel().loads(str(path)).finish()
    return xl.calculate()


def find(sol, sheet, cell, book):
    key = f"'[{book}]{sheet.upper()}'!{cell.upper()}"
    for k, v in sol.items():
        if k.upper().endswith(f"]{sheet.upper()}'!{cell.upper()}"):
            try:
                return float(v.value[0, 0])
            except Exception:
                try:
                    return float(v.value)
                except Exception:
                    return None
    return None


def label_row(path, sheet, label):
    ws = load_workbook(path)[sheet]
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == label:
            return row[0].row
    return None


def main():
    res = json.loads((DATA / "results.json").read_text())
    failures, checks = [], 0

    # Alphabet is a separate PM-ready implementation.  Verify timing, terminal,
    # shares, WACC, scenarios and decision outputs—not only the base DCF.
    ticker, slug = "GOOGL", "alphabet"
    path = OUT / slug / "model.xlsx"
    book = path.name
    pm = alphabet_pm.build_data(write=False)
    print(f"\n{'=' * 66}\n{ticker}: recalculating PM-ready {path.name}\n{'=' * 66}")
    sol = solve(path)
    expected = {
        ("Valuation", "Value per share"): pm["dcf"]["value_per_share"],
        ("Valuation", "Enterprise value"): pm["dcf"]["enterprise_value"],
        ("Valuation", "Equity value"): pm["dcf"]["equity_value"],
        ("Valuation", "Terminal value / EV"): pm["dcf"]["tv_pct_ev"],
        ("WACC", "WACC"): pm["wacc"]["wacc"],
        ("Share Bridge", "Point-in-time diluted shares"): pm["share_bridge"]["total"],
        ("Decision", "Probability-weighted value"): pm["scenarios"]["expected_value"],
    }
    for (sheet_name, label), exp in expected.items():
        row = label_row(path, sheet_name, label)
        got = find(sol, sheet_name, f"B{row}", book)
        checks += 1
        tol = TOL if abs(exp) < 1000 else max(abs(exp) * 1e-8, 1.0)
        ok = got is not None and abs(got - exp) <= tol
        print(f"  {sheet_name + ' / ' + label:<42} {'ok' if ok else 'MISMATCH'}")
        if not ok:
            failures.append(f"GOOGL {sheet_name}/{label}: workbook {got} vs python {exp}")

    # Every explicit-period FCFF and scenario target must tie.
    fcff_row = label_row(path, "Valuation", "FCFF")
    for i, exp_row in enumerate(pm["dcf"]["rows"]):
        cell = f"{chr(ord('B') + i)}{fcff_row}"
        got = find(sol, "Valuation", cell, book)
        checks += 1
        if got is None or abs(got - exp_row["fcff"]) > max(abs(exp_row["fcff"]) * 1e-8, 1.0):
            failures.append(f"GOOGL FCFF {exp_row['label']}: workbook {got} vs python {exp_row['fcff']}")
    for i, exp_row in enumerate(pm["scenarios"]["rows"], 5):
        got = find(sol, "Scenarios", f"G{i}", book)
        checks += 1
        if got is None or abs(got - exp_row["target"]) > TOL:
            failures.append(f"GOOGL scenario {exp_row['key']}: workbook {got} vs python {exp_row['target']}")

    # Formula and architecture controls catch the exact seams that previously
    # passed by construction.
    wb = load_workbook(path, data_only=False)
    required_order = ["Cover", "Sources", "Drivers", "Scenarios", "WACC", "Share Bridge",
                      "Revenue", "Depreciation", "Valuation", "Variant", "Sensitivities",
                      "Decision", "Checks", "Notes"]
    checks += 1
    if wb.sheetnames != required_order:
        failures.append(f"GOOGL sheet order {wb.sheetnames} != {required_order}")
    for cell in [wb["Depreciation"].cell(17, c) for c in range(2, 24)]:
        checks += 1
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            failures.append(f"GOOGL Depreciation {cell.coordinate} is not a formula")
    for cell in ("F5", "G5", "F6", "G6", "F7", "G7"):
        checks += 1
        if not str(wb["Scenarios"][cell].value).startswith("="):
            failures.append(f"GOOGL Scenarios {cell} is not a formula")
    for cell in ("B5", "C5", "B6", "C6", "B7", "C7", "B8", "C8"):
        checks += 1
        if not str(wb["Decision"][cell].value).startswith("="):
            failures.append(f"GOOGL Decision {cell} is not formula-linked")
    center = find(sol, "Sensitivities", "D8", book)
    checks += 1
    if center is None or abs(center - pm["dcf"]["value_per_share"]) > TOL:
        failures.append(f"GOOGL sensitivity center {center} vs base DCF {pm['dcf']['value_per_share']}")
    memo_text = (OUT / "alphabet" / "memo.html").read_text()
    for forbidden in ("<b>SHORT</b>", "Kelly f", "Quarter-Kelly", "5.00% of NAV", "$84.75bn", "$45bn of notes"):
        checks += 1
        if forbidden in memo_text:
            failures.append(f"GOOGL memo retains forbidden claim: {forbidden}")
    checks += 1
    if any(x[1] not in ("CLEARED", "PASS") for x in alphabet_pm.IMPLEMENTATION_GATES) and wb["Decision"]["B11"].value != 0:
        failures.append("GOOGL has an uncleared implementation gate but non-zero size")
    print("  formula architecture and PM gates          ", "ok" if not failures else "review")

    # Nvidia continues to use the generic two-company engine.
    for ticker, slug in (("NVDA", "nvidia"),):
        path = OUT / slug / "model.xlsx"
        book = path.name
        print(f"\n{'=' * 66}\n{ticker}: recalculating {path.name}\n{'=' * 66}")
        sol = solve(path)

        expected = {
            "Value per share": res[ticker]["scenarios"]["base"]["value_per_share"],
            "Enterprise value": res[ticker]["scenarios"]["base"]["enterprise_value"],
            "Equity value": res[ticker]["scenarios"]["base"]["equity_value"],
            "Terminal value % of EV": res[ticker]["scenarios"]["base"]["tv_pct_ev"],
        }
        for label, exp in expected.items():
            r = label_row(path, "Valuation", label)
            got = find(sol, "Valuation", f"B{r}", book)
            checks += 1
            if got is None:
                failures.append(f"{ticker} {label}: workbook produced no value")
                print(f"  {label:<26} MISSING")
                continue
            tol = TOL if "share" in label else max(abs(exp) * 1e-6, 1e-6)
            ok = abs(got - exp) <= tol
            if not ok:
                failures.append(f"{ticker} {label}: workbook {got:,.6f} vs python {exp:,.6f}")
            flag = "ok" if ok else "MISMATCH"
            if abs(exp) > 1000:
                print(f"  {label:<26} workbook {got / 1e9:>12,.3f}bn   python {exp / 1e9:>12,.3f}bn   {flag}")
            else:
                print(f"  {label:<26} workbook {got:>12,.4f}     python {exp:>12,.4f}     {flag}")

        # every forecast-year FCFF must tie as well, not just the headline
        r = label_row(path, "Valuation", "FCFF")
        for i in range(6):
            col = chr(ord("B") + i)
            got = find(sol, "Valuation", f"{col}{r}", book)
            exp = res[ticker]["scenarios"]["base"]["rows"][i]["fcff"]
            checks += 1
            if got is None or abs(got - exp) > max(abs(exp) * 1e-6, 1.0):
                failures.append(f"{ticker} FCFF year {i + 1}: workbook {got} vs python {exp:,.2f}")
        print(f"  {'FCFF, all 6 years':<26} {'ok' if not any('FCFF' in f for f in failures) else 'MISMATCH'}")

    print(f"\n{'=' * 66}")
    if failures:
        print(f"FAILED — {len(failures)} of {checks} checks did not tie:")
        for f in failures:
            print("  -", f)
        return 1
    print(f"PASSED — all {checks} checks tie between the workbook and the Python model.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
