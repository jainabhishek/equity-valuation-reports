#!/usr/bin/env python3
"""Build two screen-grade equity valuation reports and linked DCF dashboards."""

from __future__ import annotations

import argparse
import html
import json
import math
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
COMPANIES = ("alphabet", "nvidia")
SCENARIOS = ("downside", "base", "upside")

NAVY = "0B1739"
INK = "172033"
BLUE = "2864DC"
TEAL = "0A8F86"
RED = "C9495E"
PALE_BLUE = "EAF1FF"
PALE_TEAL = "E8F6F4"
PALE_RED = "FBECEF"
PALE_GRAY = "F3F6FA"
WHITE = "FFFFFF"
GRID = "D8E0EA"


CONTENT = {
    "alphabet": {
        "short_name": "Alphabet",
        "ticker": "GOOGL",
        "eyebrow": "Digital advertising · Cloud · AI infrastructure",
        "stance": "Valuation is the problem, not the franchise",
        "verdict": (
            "Alphabet remains a high-quality compounder, but the current price requires an operating outcome "
            "well beyond this base case. The model gives full credit to reported cash and securities, assumes "
            "Search stays resilient, Cloud keeps outgrowing the group, and AI capex normalizes. Even then, the "
            "base value is materially below the market. The cleanest reason to wait is cash conversion: Q2 capex "
            "exceeded operating cash flow, while reported net income was dominated by non-operating equity gains."
        ),
        "callout": (
            "The market-implied case requires roughly 25% annualized revenue growth through 2032 while keeping "
            "base margins, reinvestment, WACC, and terminal growth unchanged. That is not impossible, but it is "
            "closer to a sustained AI-platform breakout than a mature mega-cap base case."
        ),
        "metrics": [
            ("Q2 revenue", "$119.8bn", "+24% YoY"),
            ("Google Cloud", "$24.8bn", "+82% YoY"),
            ("Cloud operating margin", "35.6%", "$8.8bn operating income"),
            ("TTM free cash flow", "$53.3bn", "11.9% margin"),
            ("TTM capex", "$132.4bn", "29.7% of revenue"),
            ("Cash + marketable securities", "$242.5bn", "June 30, 2026"),
        ],
        "sections": [
            (
                "Search is still growing, but AI changes the unit economics",
                "Google Services produced $94.5bn of Q2 revenue, up 15%, and Search & other grew 17% to $63.3bn. "
                "The near-term evidence does not show a collapsing search franchise. The harder valuation question "
                "is whether AI answer formats expand query volume and commercial intent faster than they raise inference "
                "costs or displace high-value clicks. The base case assumes resilience—not a step-change in monetization."
            ),
            (
                "Cloud is now the clearest incremental value engine",
                "Google Cloud revenue rose 82% to $24.8bn and operating income more than tripled to $8.8bn. The reported "
                "35.6% segment margin demonstrates that Cloud can fund part of the infrastructure cycle rather than remain "
                "a perpetual drag. The upside case depends on Cloud sustaining elevated growth while utilization and TPU "
                "sales absorb the current data-center build."
            ),
            (
                "Reported earnings overstate recurring economics",
                "Q2 other income was $98.0bn, including $99.0bn of equity-security gains. That lifted quarterly net income "
                "to $112.1bn, but it does not represent recurring operating profit. The DCF therefore starts from EBIT and "
                "cash reinvestment, not the headline P/E. This is why Alphabet can look inexpensive on trailing earnings "
                "while still screen as expensive on normalized free cash flow."
            ),
            (
                "AI capacity is a real claim on cash",
                "Alphabet spent $80.6bn on capex in the first half of 2026 and $44.9bn in Q2 alone, versus $39.1bn of Q2 "
                "operating cash flow. The company also raised $49.6bn through common and mandatory-convertible preferred "
                "equity in June. The base case assumes capex falls from 28% of revenue in 2027 to 11% by 2032; failure to "
                "normalize is the single largest cash-conversion risk."
            ),
        ],
        "catalysts": [
            "Cloud growth and margin remain structurally above the group as TPU systems and AI services scale.",
            "Search AI experiences increase commercial queries without materially weakening ad yield.",
            "Capex peaks, depreciation catches up, and free cash flow reconnects with operating income.",
            "Waymo or other non-core assets create monetizable value not captured in the operating DCF.",
        ],
        "risks": [
            "AI interfaces erode Search monetization or materially increase inference cost per query.",
            "Data-center and power commitments lock in spending before utilization and returns are proven.",
            "Regulatory remedies impair distribution, advertising practices, or platform economics.",
            "Equity-security gains reverse, obscuring the gap between headline and normalized earnings.",
            "Preferred/common issuance creates more dilution than the current bridge captures.",
        ],
        "questions": [
            "When should capex fall below operating cash flow on a sustained basis?",
            "How much of Cloud's acceleration is capacity-constrained backlog versus durable end demand?",
            "What is the ad-revenue and cost-per-query profile of AI Overviews and agentic Search?",
            "How much of the $242.5bn liquidity balance is truly excess after committed AI infrastructure funding?",
        ],
        "sources": [
            ("Alphabet Q2 2026 Form 10-Q", "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000071/goog-20260630.htm"),
            ("Alphabet Q2 2026 earnings release", "https://s206.q4cdn.com/479360582/files/doc_financials/2026/q2/2026q2-alphabet-earnings-release.pdf"),
            ("GOOGL market price and TTM financials", "https://stockanalysis.com/stocks/googl/financials/"),
            ("U.S. Treasury daily yield curve", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?field_tdr_date_value=2026&type=daily_treasury_yield_curve"),
        ],
    },
    "nvidia": {
        "short_name": "Nvidia",
        "ticker": "NVDA",
        "eyebrow": "Accelerated computing · Networking · AI infrastructure",
        "stance": "Exceptional business; the price demands continued exceptionalism",
        "verdict": (
            "Nvidia's operating performance still supports the strongest fundamental case in large-cap AI: Q1 revenue "
            "grew 85%, Data Center grew 92%, gross margin held near 75%, and operating cash flow reached $50.3bn. The "
            "valuation, however, capitalizes a long period of hyperscaler and sovereign AI spending with only gradual "
            "margin normalization. The base DCF lands below the market; the upside case clears it, making this a business "
            "quality versus expectation-risk decision rather than a broken-thesis call."
        ),
        "callout": (
            "The market-implied case requires roughly 29% annualized revenue growth through 2032—about nine percentage "
            "points above the base growth path each year—while holding base margins, working capital, WACC, and terminal "
            "growth constant. That hurdle is demanding but meaningfully more attainable than Alphabet's implied case."
        ),
        "metrics": [
            ("Q1 FY2027 revenue", "$81.6bn", "+85% YoY"),
            ("Data Center revenue", "$75.2bn", "+92% YoY"),
            ("Data Center networking", "$14.8bn", "+199% YoY"),
            ("GAAP gross margin", "74.9%", "Q1 FY2027"),
            ("Q1 operating cash flow", "$50.3bn", "61.7% of revenue"),
            ("Cash + securities", "$80.6bn", "April 26, 2026"),
        ],
        "sections": [
            (
                "The platform is broader than GPUs",
                "Data Center revenue reached $75.2bn in Q1, including $60.4bn of compute and $14.8bn of networking. "
                "Networking growth of 199% shows that Nvidia is monetizing the fabric around accelerated compute, not "
                "only the accelerator. NVLink, Ethernet, InfiniBand, systems, and software deepen switching costs and "
                "raise the economic value of each architecture cycle."
            ),
            (
                "Margins remain extraordinary, but normalization matters",
                "Q1 GAAP gross margin was 74.9% and TTM operating margin was roughly 64%. The base case deliberately fades "
                "EBIT margin from 63% in 2027 to 54% by 2032 as system content rises, competition expands, and customers "
                "gain bargaining power. Small changes to that terminal margin produce large changes in value because "
                "Nvidia's current economics are far above mature semiconductor norms."
            ),
            (
                "Demand visibility is strong and concentrated",
                "Nvidia guided Q2 revenue to $91bn plus or minus 2%, excluding China data-center compute revenue. At the "
                "same time, the FY2026 10-K disclosed that one direct customer represented 22% of revenue and another 14%. "
                "Large customers validate demand, but they also concentrate negotiating power and are the same firms most "
                "able to fund custom accelerators."
            ),
            (
                "Cash conversion is excellent; ecosystem investment complicates the bridge",
                "TTM free cash flow was about $119.1bn and Q1 capex was only $1.8bn because Nvidia remains fabless. Cash, "
                "marketable debt securities, and marketable equity securities totaled $80.6bn. The company is also making "
                "large strategic investments; Q1 net income included $15.9bn of equity-security gains, so this report "
                "anchors on operating income and FCFF rather than headline earnings."
            ),
        ],
        "catalysts": [
            "Q2 and subsequent revenue exceed the $91bn guide as Blackwell and networking supply ramps.",
            "Rubin extends architecture leadership and shortens customer payback periods.",
            "Networking, software, inference, and sovereign AI reduce dependence on training-only demand.",
            "Gross margin remains above the modeled fade despite a rising full-system mix.",
        ],
        "risks": [
            "Hyperscaler capex growth slows after the current AI factory buildout.",
            "Custom ASICs or competing accelerators capture high-volume workloads.",
            "Export controls eliminate additional demand or force lower-value product configurations.",
            "Foundry, packaging, memory, power, and networking constraints disrupt product ramps.",
            "Customer concentration and architecture transitions create order volatility or inventory charges.",
        ],
        "questions": [
            "What portion of current demand is limited by supply versus customer capital budgets?",
            "How should investors underwrite gross margin as full systems and networking become a larger mix?",
            "How durable is customer ROI if AI model costs and inference pricing fall faster than usage grows?",
            "Can software and services become large enough to offset eventual hardware-cycle normalization?",
        ],
        "sources": [
            ("Nvidia Q1 FY2027 Form 10-Q", "https://www.sec.gov/Archives/edgar/data/1045810/000104581026000052/nvda-20260426.htm"),
            ("Nvidia FY2026 Form 10-K", "https://www.sec.gov/Archives/edgar/data/1045810/000104581026000021/nvda-20260125.htm"),
            ("Nvidia Q1 FY2027 earnings release", "https://nvidianews.nvidia.com/news/nvidia-announces-financial-results-for-first-quarter-fiscal-2027"),
            ("NVDA market price and TTM financials", "https://stockanalysis.com/stocks/nvda/financials/"),
            ("U.S. Treasury daily yield curve", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?field_tdr_date_value=2026&type=daily_treasury_yield_curve"),
        ],
    },
}


def load_plan(company: str) -> dict:
    return json.loads((ROOT / company / "plan.json").read_text())


def dcf(plan: dict, scenario: str = "base", growth_delta: float = 0.0) -> dict:
    case = plan["scenarios"][scenario]
    horizon = plan["timeline"]["horizon_years"]
    growth = [float(x) + growth_delta for x in case["revenue_growth"]]
    ebit_margin = [float(x) for x in case["ebit_margin"]]
    taxes = [float(x) for x in case["tax_rate"]]
    da_pct = [float(x) for x in case["da_percent_revenue"]]
    capex_pct = [float(x) for x in case["capex_percent_revenue"]]
    nwc_pct = [float(x) for x in case["nwc_percent_revenue"]]
    w = plan["wacc"]
    cost_equity = w["risk_free_rate"] + w["beta"] * w["equity_risk_premium"] + w.get("size_premium", 0) + w.get("company_specific_premium", 0) + w.get("country_risk_premium", 0)
    after_tax_debt = w["pre_tax_cost_of_debt"] * (1 - w["marginal_tax_rate"])
    wacc = w["target_equity_pct"] * cost_equity + w["target_debt_pct"] * after_tax_debt + w.get("preferred_pct", 0) * w.get("pre_tax_cost_of_preferred", 0) + case["wacc_adjustment"]
    revenue = float(plan["historicals"]["revenue"])
    previous_nwc = float(plan["historicals"].get("net_working_capital", 0))
    rows = []
    pv_explicit = 0.0
    for i in range(horizon):
        revenue *= 1 + growth[i]
        ebit = revenue * ebit_margin[i]
        nopat = ebit * (1 - taxes[i])
        da = revenue * da_pct[i]
        capex = revenue * capex_pct[i]
        nwc = revenue * nwc_pct[i]
        change_nwc = nwc - previous_nwc
        fcff = nopat + da - capex - change_nwc
        pv = fcff / ((1 + wacc) ** (i + 0.5))
        rows.append({"year": 2027 + i, "growth": growth[i], "revenue": revenue, "ebit_margin": ebit_margin[i], "ebit": ebit, "fcff": fcff, "pv": pv})
        pv_explicit += pv
        previous_nwc = nwc
    terminal_growth = float(case["terminal_growth_rate"])
    terminal_fcf = rows[-1]["fcff"] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth)
    pv_terminal = terminal_value / ((1 + wacc) ** horizon)
    enterprise_value = pv_explicit + pv_terminal
    b = plan["ev_to_equity_bridge"]
    equity_value = enterprise_value + b["cash"] + b.get("non_operating_assets", 0) + b.get("associates", 0) - b["debt"] - b.get("leases", 0) - b.get("minorities", 0) - b.get("pensions", 0) - b.get("preferred_stock", 0) - b.get("options", 0) - b.get("other_debt_like_items", 0)
    value_per_share = equity_value / b["diluted_shares"]
    return {
        "scenario": scenario,
        "rows": rows,
        "wacc": wacc,
        "terminal_growth": terminal_growth,
        "pv_explicit": pv_explicit,
        "pv_terminal": pv_terminal,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "value_per_share": value_per_share,
        "tv_percent_ev": pv_terminal / enterprise_value,
        "revenue_cagr": math.prod(1 + x for x in growth) ** (1 / horizon) - 1,
    }


def reverse_dcf(plan: dict) -> dict:
    target = float(plan["market"]["current_share_price"])
    low, high = -0.20, 0.40
    for _ in range(120):
        mid = (low + high) / 2
        if dcf(plan, "base", mid)["value_per_share"] < target:
            low = mid
        else:
            high = mid
    delta = (low + high) / 2
    result = dcf(plan, "base", delta)
    result["growth_delta"] = delta
    return result


def set_cell(ws, cell, value, *, fill=None, color=INK, bold=False, size=10, number_format=None, align=None):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Aptos", size=size, color=color, bold=bold)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if number_format:
        c.number_format = number_format
    c.alignment = align or Alignment(vertical="center")
    return c


def engine_block(ws, plan, scenario, start_col, reverse=False):
    cp_rows = {
        "downside": {"growth": 22, "ebit": 32, "da": 37, "capex": 42, "nwc": 47, "tax": 52},
        "base": {"growth": 21, "ebit": 31, "da": 36, "capex": 41, "nwc": 46, "tax": 51},
        "upside": {"growth": 23, "ebit": 33, "da": 38, "capex": 43, "nwc": 48, "tax": 53},
    }[scenario]
    name = "Reverse DCF" if reverse else scenario.title()
    start = get_column_letter(start_col)
    end = get_column_letter(start_col + 5)
    ws.merge_cells(f"{start}24:{end}24")
    set_cell(ws, f"{start}24", name, fill=NAVY, color=WHITE, bold=True, size=11, align=Alignment(horizontal="center"))
    labels = {
        26: "Revenue growth", 27: "Revenue", 28: "EBIT margin", 29: "EBIT", 30: "Tax rate", 31: "NOPAT",
        32: "D&A / revenue", 33: "D&A", 34: "Capex / revenue", 35: "Capex", 36: "NWC / revenue", 37: "NWC",
        38: "Change in NWC", 39: "FCFF", 40: "Discount period", 41: "PV of FCFF"
    }
    for row, label in labels.items():
        set_cell(ws, f"A{row}", label, color="5B667A", size=9)
    for i in range(6):
        col = get_column_letter(start_col + i)
        cp_col = get_column_letter(2 + i)
        set_cell(ws, f"{col}25", f"='Control Panel'!{cp_col}$18", fill=PALE_GRAY, bold=True, align=Alignment(horizontal="center"))
        growth_formula = f"='Control Panel'!{cp_col}${cp_rows['growth']}"
        if reverse:
            growth_formula += "+$C$18"
        set_cell(ws, f"{col}26", growth_formula, number_format="0.0%")
        if i == 0:
            set_cell(ws, f"{col}27", f"='Historical Financials'!$D$6*(1+{col}26)", number_format="$#,##0")
        else:
            prev = get_column_letter(start_col + i - 1)
            set_cell(ws, f"{col}27", f"={prev}27*(1+{col}26)", number_format="$#,##0")
        set_cell(ws, f"{col}28", f"='Control Panel'!{cp_col}${cp_rows['ebit']}", number_format="0.0%")
        set_cell(ws, f"{col}29", f"={col}27*{col}28", number_format="$#,##0")
        set_cell(ws, f"{col}30", f"='Control Panel'!{cp_col}${cp_rows['tax']}", number_format="0.0%")
        set_cell(ws, f"{col}31", f"={col}29*(1-{col}30)", number_format="$#,##0")
        set_cell(ws, f"{col}32", f"='Control Panel'!{cp_col}${cp_rows['da']}", number_format="0.0%")
        set_cell(ws, f"{col}33", f"={col}27*{col}32", number_format="$#,##0")
        set_cell(ws, f"{col}34", f"='Control Panel'!{cp_col}${cp_rows['capex']}", number_format="0.0%")
        set_cell(ws, f"{col}35", f"={col}27*{col}34", number_format="$#,##0")
        set_cell(ws, f"{col}36", f"='Control Panel'!{cp_col}${cp_rows['nwc']}", number_format="0.0%")
        set_cell(ws, f"{col}37", f"={col}27*{col}36", number_format="$#,##0")
        if i == 0:
            set_cell(ws, f"{col}38", f"={col}37-'Historical Financials'!$D$18", number_format="$#,##0")
        else:
            prev = get_column_letter(start_col + i - 1)
            set_cell(ws, f"{col}38", f"={col}37-{prev}37", number_format="$#,##0")
        set_cell(ws, f"{col}39", f"={col}31+{col}33-{col}35-{col}38", number_format="$#,##0")
        set_cell(ws, f"{col}40", i + 0.5, number_format="0.0x")
        set_cell(ws, f"{col}41", f"={col}39/(1+{start}$43)^{col}40", number_format="$#,##0")
    adjustment = 0 if reverse else plan["scenarios"][scenario]["wacc_adjustment"]
    terminal_growth = plan["scenarios"][scenario]["terminal_growth_rate"] if not reverse else plan["scenarios"]["base"]["terminal_growth_rate"]
    metrics = [
        (43, "WACC", f"='Control Panel'!$B$77+{adjustment}", "0.0%"),
        (44, "Terminal growth", terminal_growth, "0.0%"),
        (45, "Terminal FCFF", f"={end}39*(1+{start}44)", "$#,##0"),
        (46, "Terminal value", f"={start}45/({start}43-{start}44)", "$#,##0"),
        (47, "PV of terminal value", f"={start}46/(1+{start}43)^6", "$#,##0"),
        (48, "PV of explicit FCFF", f"=SUM({start}41:{end}41)", "$#,##0"),
        (49, "Enterprise value", f"={start}47+{start}48", "$#,##0"),
        (50, "Net debt", "='Control Panel'!$B$85", "$#,##0"),
        (51, "Minority interest", "='Control Panel'!$B$87", "$#,##0"),
        (52, "Preferred stock", "='Control Panel'!$B$88", "$#,##0"),
        (53, "Non-operating assets", "='Control Panel'!$B$89", "$#,##0"),
        (54, "Other equity adjustments", "='Control Panel'!$B$90", "$#,##0"),
        (55, "Equity value", f"={start}49-{start}50-{start}51-{start}52+{start}53+{start}54", "$#,##0"),
        (56, "Diluted shares", "='Control Panel'!$B$86", "#,##0"),
        (57, "Value per share", f"={start}55/{start}56", "$0.00"),
        (58, "Current share price", "='Control Panel'!$B$12", "$0.00"),
        (59, "Premium / (discount)", f"={start}57/{start}58-1", "0.0%"),
        (60, "Terminal value / EV", f"={start}47/{start}49", "0.0%"),
    ]
    for row, label, value, fmt in metrics:
        set_cell(ws, f"A{row}", label, color="5B667A", size=9)
        set_cell(ws, f"{start}{row}", value, fill=PALE_GRAY if row not in (57, 59) else PALE_BLUE, bold=row in (49, 55, 57), number_format=fmt)
    return start


def prepare_workbook(company: str):
    directory = ROOT / company
    plan = load_plan(company)
    reverse = reverse_dcf(plan)
    path = directory / "banker_formula_workbook.xlsx"
    wb = load_workbook(path)
    if "Dashboard" in wb.sheetnames:
        wb.remove(wb["Dashboard"])
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B24"
    ws.merge_cells("A1:J1")
    set_cell(ws, "A1", f"{CONTENT[company]['short_name']} — Intrinsic Value Dashboard", fill=NAVY, color=WHITE, bold=True, size=18, align=Alignment(horizontal="left"))
    ws.row_dimensions[1].height = 34
    set_cell(ws, "A3", "Calculation integrity", fill=PALE_TEAL, color=TEAL, bold=True)
    set_cell(ws, "B3", "Formula workbook; recalculated and error-scanned in the run log", fill=PALE_TEAL, color=INK)
    set_cell(ws, "A4", "Decision readiness", fill=PALE_RED, color=RED, bold=True)
    set_cell(ws, "B4", "SCREEN-GRADE — public actuals plus analyst forecast/WACC assumptions", fill=PALE_RED, color=INK)
    set_cell(ws, "A6", "Valuation date", color="5B667A")
    set_cell(ws, "B6", plan["meta"]["valuation_date"], bold=True)
    set_cell(ws, "A7", "Current price", color="5B667A")
    set_cell(ws, "B7", "='Control Panel'!$B$12", bold=True, size=14, number_format="$0.00")
    headers = ("Downside", "Base", "Upside")
    fills = (PALE_RED, PALE_BLUE, PALE_TEAL)
    for idx, (header, fill) in enumerate(zip(headers, fills), start=2):
        set_cell(ws, f"{get_column_letter(idx)}8", header, fill=fill, bold=True, align=Alignment(horizontal="center"))
    scenario_starts = {}
    for scenario, col in zip(SCENARIOS, (2, 10, 18)):
        scenario_starts[scenario] = engine_block(ws, plan, scenario, col)
    reverse_start = engine_block(ws, plan, "base", 26, reverse=True)
    summary_rows = [(9, "WACC", 43, "0.0%"), (10, "Terminal growth", 44, "0.0%"), (11, "Revenue CAGR", None, "0.0%"), (12, "Enterprise value", 49, "$#,##0"), (13, "Value per share", 57, "$0.00"), (14, "Premium / (discount)", 59, "0.0%"), (15, "Terminal value / EV", 60, "0.0%")]
    for row, label, source_row, fmt in summary_rows:
        set_cell(ws, f"A{row}", label, color="5B667A", bold=row in (13, 14))
        for idx, scenario in enumerate(SCENARIOS, start=2):
            start = scenario_starts[scenario]
            if row == 11:
                growth_cells = [f"(1+{get_column_letter(ord(start) - 64 + i)}26)" for i in range(6)]
                value = "=(" + "*".join(growth_cells) + ")^(1/6)-1"
            else:
                value = f"={start}{source_row}"
            set_cell(ws, f"{get_column_letter(idx)}{row}", value, fill=PALE_GRAY if row not in (13, 14) else fills[idx - 2], bold=row in (13, 14), number_format=fmt, align=Alignment(horizontal="center"))
    set_cell(ws, "A17", "Reverse DCF", fill=NAVY, color=WHITE, bold=True)
    set_cell(ws, "A18", "Uniform growth uplift vs base", color="5B667A")
    set_cell(ws, "C18", reverse["growth_delta"], fill=PALE_BLUE, bold=True, number_format="0.0%")
    set_cell(ws, "A19", "Market-implied six-year revenue CAGR", color="5B667A")
    reverse_growth_cells = [f"(1+{get_column_letter(26 + i)}26)" for i in range(6)]
    set_cell(ws, "C19", "=(" + "*".join(reverse_growth_cells) + ")^(1/6)-1", fill=PALE_BLUE, bold=True, number_format="0.0%")
    set_cell(ws, "A20", "Reverse-DCF value per share", color="5B667A")
    set_cell(ws, "C20", f"={reverse_start}57", fill=PALE_BLUE, bold=True, number_format="$0.00")
    set_cell(ws, "F6", "Core method", color="5B667A")
    set_cell(ws, "G6", "6-year FCFF, mid-year convention, Gordon growth", bold=True)
    set_cell(ws, "F7", "Current price source", color="5B667A")
    set_cell(ws, "G7", plan["market"]["source"], color=BLUE)
    for col in range(1, 32):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    for row in range(8, 61):
        ws.row_dimensions[row].height = 18
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows(min_row=8, max_row=60, min_col=1, max_col=31):
        for cell in row:
            if cell.value is not None:
                cell.border = Border(bottom=thin)
    ws.auto_filter.ref = "A8:D15"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_area = "A1:J20"
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


def money(value: float, decimals=0) -> str:
    sign = "-" if value < 0 else ""
    value = abs(value)
    if value >= 1_000_000:
        return f"{sign}${value / 1_000_000:.2f}tn"
    if value >= 1_000:
        return f"{sign}${value / 1_000:.1f}bn"
    return f"{sign}${value:.{decimals}f}mm"


def pct(value: float, decimals=1) -> str:
    return f"{value * 100:.{decimals}f}%"


def html_report(company: str, results: dict, reverse: dict, formula_count: int, error_count: int) -> str:
    content = CONTENT[company]
    plan = load_plan(company)
    current = plan["market"]["current_share_price"]
    base = results["base"]
    discount = base["value_per_share"] / current - 1
    max_value = max(current, *(x["value_per_share"] for x in results.values()))
    scenario_rows = "".join(
        f"<tr><td><strong>{s.title()}</strong><span>{html.escape(plan['scenarios'][s]['description'])}</span></td>"
        f"<td>{pct(r['revenue_cagr'])}</td><td>{pct(r['rows'][-1]['ebit_margin'])}</td><td>{pct(r['wacc'])}</td>"
        f"<td>{pct(r['terminal_growth'])}</td><td>{pct(r['tv_percent_ev'])}</td><td><strong>${r['value_per_share']:.2f}</strong></td></tr>"
        for s, r in ((s, results[s]) for s in SCENARIOS)
    )
    forecast_rows = "".join(
        f"<tr><td>{r['year']}E</td><td>{money(r['revenue'])}</td><td>{pct(r['growth'])}</td><td>{pct(r['ebit_margin'])}</td><td>{money(r['fcff'])}</td></tr>"
        for r in base["rows"]
    )
    metric_cards = "".join(f"<div class='metric'><span>{html.escape(k)}</span><strong>{v}</strong><small>{html.escape(note)}</small></div>" for k, v, note in content["metrics"])
    deep_sections = "".join(f"<article><h3>{html.escape(title)}</h3><p>{html.escape(body)}</p></article>" for title, body in content["sections"])
    catalysts = "".join(f"<li>{html.escape(x)}</li>" for x in content["catalysts"])
    risks = "".join(f"<li>{html.escape(x)}</li>" for x in content["risks"])
    questions = "".join(f"<li>{html.escape(x)}</li>" for x in content["questions"])
    sources = "".join(f"<li><a href='{url}'>{html.escape(label)}</a></li>" for label, url in content["sources"])
    bars = "".join(
        f"<div class='bar-row'><span>{s.title()}</span><div class='track'><i class='{s}' style='width:{max(5, results[s]['value_per_share']/max_value*100):.1f}%'></i></div><b>${results[s]['value_per_share']:.2f}</b></div>"
        for s in SCENARIOS
    ) + f"<div class='bar-row market'><span>Market</span><div class='track'><i style='width:{current/max_value*100:.1f}%'></i></div><b>${current:.2f}</b></div>"
    integrity = "PASS" if error_count == 0 else "REVIEW"
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><link rel="icon" href="data:,">
<title>{content['short_name']} intrinsic value report — August 4, 2026</title>
<style>
:root{{--navy:#0b1739;--ink:#172033;--muted:#667085;--blue:#2864dc;--teal:#0a8f86;--red:#c9495e;--pale:#f3f6fa;--line:#d8e0ea;--white:#fff}}
*{{box-sizing:border-box}}body{{margin:0;background:#edf2f7;color:var(--ink);font:15px/1.55 Inter,Aptos,system-ui,-apple-system,sans-serif}}a{{color:var(--blue)}}
.page{{max-width:1180px;margin:28px auto;background:#fff;box-shadow:0 18px 55px rgba(11,23,57,.12)}}
header{{background:linear-gradient(135deg,#0b1739 0%,#142958 70%,#183668 100%);color:#fff;padding:54px 64px 46px}}
.eyebrow{{text-transform:uppercase;letter-spacing:.14em;font-size:12px;color:#a8c4ff;font-weight:800}}h1{{font-size:46px;line-height:1.05;margin:14px 0 16px;letter-spacing:-.035em}}header p{{max-width:800px;color:#dce7ff;font-size:18px;margin:0}}
.stamp{{display:flex;gap:10px;flex-wrap:wrap;margin-top:26px}}.chip{{border:1px solid rgba(255,255,255,.28);border-radius:999px;padding:7px 12px;font-size:12px;font-weight:750;background:rgba(255,255,255,.08)}}
main{{padding:42px 64px 60px}}.hero{{display:grid;grid-template-columns:1.05fr .95fr;gap:28px}}.card{{border:1px solid var(--line);border-radius:16px;padding:24px;background:#fff}}.verdict{{border-top:5px solid var(--blue)}}
.verdict h2{{font-size:26px;margin:0 0 12px;line-height:1.2}}.verdict p{{font-size:16px;margin:0;color:#344054}}.value-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;margin-top:22px}}
.value-grid div{{background:var(--pale);padding:16px;border-radius:12px}}.value-grid span,.metric span{{display:block;color:var(--muted);font-size:12px;text-transform:uppercase;letter-spacing:.05em;font-weight:750}}.value-grid strong{{display:block;font-size:27px;margin-top:3px}}.negative{{color:var(--red)}}
.integrity{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin:26px 0}}.integrity .card{{padding:18px 20px}}.integrity h3{{font-size:13px;text-transform:uppercase;letter-spacing:.08em;margin:0 0 5px}}.integrity strong{{font-size:19px}}.integrity p{{margin:5px 0 0;color:var(--muted);font-size:13px}}
.calc{{border-left:5px solid var(--teal);background:#f2fbfa}}.ready{{border-left:5px solid var(--red);background:#fff6f7}}
.chart h2,.section h2{{font-size:27px;margin:0 0 18px;letter-spacing:-.02em}}.bar-row{{display:grid;grid-template-columns:82px 1fr 76px;gap:12px;align-items:center;margin:14px 0}}.bar-row span,.bar-row b{{font-size:13px}}.track{{height:18px;border-radius:5px;background:#e8edf4;overflow:hidden}}.track i{{display:block;height:100%;background:var(--blue)}}.track i.downside{{background:var(--red)}}.track i.upside{{background:var(--teal)}}.market .track i{{background:var(--navy)}}
.callout{{margin:30px 0;background:#eef4ff;border:1px solid #cbdcff;border-radius:15px;padding:22px 24px;font-size:17px}}.callout strong{{color:var(--blue)}}
.metrics{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:26px 0 42px}}.metric{{border:1px solid var(--line);border-radius:13px;padding:17px}}.metric strong{{display:block;font-size:23px;margin:4px 0}}.metric small{{color:var(--muted)}}
.section{{margin-top:46px}}.prose-grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}article{{border-top:3px solid var(--navy);padding:16px 8px 8px 0}}article h3{{font-size:18px;line-height:1.25;margin:0 0 8px}}article p{{margin:0;color:#475467}}
table{{width:100%;border-collapse:separate;border-spacing:0;border:1px solid var(--line);border-radius:12px;overflow:hidden;font-size:13px}}th{{background:var(--navy);color:#fff;text-align:right;padding:12px 11px}}th:first-child,td:first-child{{text-align:left}}td{{border-top:1px solid var(--line);padding:12px 11px;text-align:right;vertical-align:top}}td span{{display:block;max-width:390px;color:var(--muted);font-size:12px;margin-top:3px}}tbody tr:nth-child(even){{background:#fafbfd}}
.twocol{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.list-card{{border-radius:14px;padding:22px 24px;border:1px solid var(--line)}}.list-card h3{{margin:0 0 10px}}.list-card ul{{margin:0;padding-left:18px}}.list-card li{{margin:8px 0}}.catalysts{{background:#f2fbfa;border-color:#bfe5e0}}.risks{{background:#fff6f7;border-color:#f1c8d0}}
.questions{{background:var(--pale);border-radius:14px;padding:24px 28px}}.questions ol{{margin:0;padding-left:22px}}.questions li{{margin:9px 0}}
.method{{color:#475467;font-size:13px}}.sources{{columns:2;padding-left:18px}}.sources li{{break-inside:avoid;margin:8px 0}}.download{{display:inline-flex;margin-top:16px;background:var(--navy);color:#fff;text-decoration:none;border-radius:9px;padding:11px 15px;font-weight:750}}
footer{{border-top:1px solid var(--line);padding:22px 64px 34px;color:var(--muted);font-size:12px}}
@media(max-width:800px){{.page{{margin:0}}header,main,footer{{padding-left:24px;padding-right:24px}}h1{{font-size:36px}}.hero,.prose-grid,.twocol,.integrity{{grid-template-columns:1fr}}.metrics{{grid-template-columns:1fr 1fr}}table{{display:block;overflow-x:auto}}}}
@media(max-width:520px){{.metrics,.value-grid{{grid-template-columns:1fr}}}}
@media print{{body{{background:#fff}}.page{{margin:0;box-shadow:none}}a{{color:inherit}}}}
</style></head><body><div class="page">
<header><div class="eyebrow">{content['eyebrow']}</div><h1>{content['short_name']} intrinsic value report</h1><p>{content['stance']}. Standalone operating and DCF deep dive as of August 4, 2026.</p><div class="stamp"><span class="chip">{content['ticker']}</span><span class="chip">Screen-grade</span><span class="chip">6-year FCFF DCF</span><span class="chip">USD</span></div></header>
<main><section class="hero"><div class="card verdict"><h2>{content['stance']}</h2><p>{content['verdict']}</p><div class="value-grid"><div><span>Current price</span><strong data-model-cite="Dashboard!B7">${current:.2f}</strong></div><div><span>Base intrinsic value</span><strong data-model-cite="Dashboard!C13">${base['value_per_share']:.2f}</strong></div><div><span>Premium / (discount)</span><strong class="negative" data-model-cite="Dashboard!C14">{pct(discount)}</strong></div><div><span>Scenario range</span><strong data-model-cite="Dashboard!B13:D13">${results['downside']['value_per_share']:.0f}–${results['upside']['value_per_share']:.0f}</strong></div></div></div><div class="card chart"><h2>DCF value versus market</h2>{bars}</div></section>
<section class="integrity"><div class="card calc"><h3>Calculation integrity</h3><strong>{integrity}</strong><p>{formula_count:,} formulas; {error_count} spreadsheet errors after recalculation. Scenario outputs tie to the Dashboard.</p></div><div class="card ready"><h3>Decision readiness</h3><strong>SCREEN-GRADE</strong><p>Public actuals are sourced; forecasts, normalized beta, ERP, and terminal assumptions are analyst judgments.</p></div></section>
<div class="callout"><strong>What the market is pricing:</strong> {content['callout']} The workbook goal-seek uses a <strong>{pct(reverse['growth_delta'])}</strong> annual growth uplift and reaches <strong>{pct(reverse['revenue_cagr'])}</strong> implied revenue CAGR at the current price.</div>
<div class="metrics">{metric_cards}</div>
<section class="section"><h2>Business and earnings-quality deep dive</h2><div class="prose-grid">{deep_sections}</div></section>
<section class="section"><h2>Base-case operating build</h2><table><thead><tr><th>Year</th><th>Revenue</th><th>Growth</th><th>EBIT margin</th><th>FCFF</th></tr></thead><tbody>{forecast_rows}</tbody></table></section>
<section class="section"><h2>Scenario valuation</h2><table><thead><tr><th>Case</th><th>Revenue CAGR</th><th>2032 EBIT margin</th><th>WACC</th><th>Terminal growth</th><th>TV / EV</th><th>Value / share</th></tr></thead><tbody>{scenario_rows}</tbody></table></section>
<section class="section"><h2>Catalysts and risks</h2><div class="twocol"><div class="list-card catalysts"><h3>Catalysts</h3><ul>{catalysts}</ul></div><div class="list-card risks"><h3>Risks</h3><ul>{risks}</ul></div></div></section>
<section class="section"><h2>Questions that could change the valuation</h2><div class="questions"><ol>{questions}</ol></div></section>
<section class="section method"><h2>Methodology and model limits</h2><p>The report uses a six-year unlevered free-cash-flow DCF with a mid-year convention and Gordon-growth terminal value. Enterprise value is bridged to equity using the latest reported cash, securities, debt, leases, preferred stock where applicable, and diluted shares. The base WACC is {pct(base['wacc'])}; terminal growth is {pct(base['terminal_growth'])}; terminal value is {pct(base['tv_percent_ev'])} of enterprise value. Forecasts are independent analyst estimates, not management guidance or consensus. Current price is the August 4, 2026 regular-session close. This is a valuation screen, not personalized investment advice.</p><a class="download" href="banker_formula_workbook.xlsx">Open the linked DCF workbook</a></section>
<section class="section method"><h2>Sources</h2><ul class="sources">{sources}</ul></section></main>
<footer>Prepared August 4, 2026 · {content['short_name']} standalone intrinsic value report · Model outputs are linked to Dashboard cells in the companion workbook.</footer>
</div></body></html>"""


def finalize(company: str):
    directory = ROOT / company
    plan = load_plan(company)
    path = directory / "banker_formula_workbook.xlsx"
    formula_wb = load_workbook(path, data_only=False)
    value_wb = load_workbook(path, data_only=True)
    formula_count = sum(1 for ws in formula_wb for row in ws.iter_rows() for c in row if c.data_type == "f")
    errors = []
    for ws in value_wb:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append({"sheet": ws.title, "cell": cell.coordinate, "error": cell.value})
    results = {s: dcf(plan, s) for s in SCENARIOS}
    reverse = reverse_dcf(plan)
    dashboard = value_wb["Dashboard"]
    cell_map = {"downside": "B13", "base": "C13", "upside": "D13"}
    tie_out = {}
    for s, cell in cell_map.items():
        workbook_value = float(dashboard[cell].value)
        delta = workbook_value - results[s]["value_per_share"]
        tie_out[s] = {"cell": f"Dashboard!{cell}", "workbook_value": workbook_value, "independent_value": results[s]["value_per_share"], "difference": delta}
        if abs(delta) > 0.02:
            raise RuntimeError(f"{company} {s} scenario does not tie: {delta}")
    reverse_delta = float(dashboard["C20"].value) - plan["market"]["current_share_price"]
    if abs(reverse_delta) > 0.02:
        raise RuntimeError(f"{company} reverse DCF does not tie to current price: {reverse_delta}")
    (directory / "report.html").write_text(html_report(company, results, reverse, formula_count, len(errors)), encoding="utf-8")
    now = datetime.now(timezone.utc).isoformat()
    citations = []
    metrics = {
        "current_price": "B7", "downside_value_per_share": "B13", "base_value_per_share": "C13", "upside_value_per_share": "D13",
        "base_discount_to_market": "C14", "base_terminal_value_percent_ev": "C15", "reverse_growth_uplift": "C18",
        "market_implied_revenue_cagr": "C19", "reverse_dcf_value_per_share": "C20",
    }
    fws = formula_wb["Dashboard"]
    vws = value_wb["Dashboard"]
    for metric, cell in metrics.items():
        citations.append({
            "citation_id": f"model-output:dashboard:{cell.lower()}", "source_id": f"model-output:dashboard:{cell.lower()}",
            "parent_source_id": "model-output", "title": metric.replace("_", " ").title(), "short_label": f"Model: Dashboard!{cell}",
            "type": "model_cell", "quality": "model_output", "workbook_path": str(path.relative_to(ROOT.parent.parent.parent.parent)),
            "sheet": "Dashboard", "cell_or_range": cell, "cell": cell, "range": cell, "metric_name": metric,
            "value": vws[cell].value, "formula": fws[cell].value if fws[cell].data_type == "f" else "", "source_ids": ["model-output"],
            "assumption_flag": metric in {"reverse_growth_uplift"}, "tie_out_status": "tied", "generated_at": now,
            "notes": "Used by the standalone HTML report; workbook is the numerical source of truth."
        })
    (directory / "model_citations.json").write_text(json.dumps(citations, indent=2), encoding="utf-8")
    run_log_path = directory / "banker_formula_workbook_run_log.json"
    run_log = json.loads(run_log_path.read_text())
    run_log.update({
        "model_status": "screen-grade", "first_visible_sheet": formula_wb.sheetnames[0],
        "formula_error_scan": {"formula_count": formula_count, "error_count": len(errors), "errors": errors, "recalculation_engine": "LibreOffice 26.8"},
        "scenario_tie_out": tie_out, "reverse_dcf_tie_out": {"cell": "Dashboard!C20", "difference_to_current_price": reverse_delta},
        "html_report": str(directory / "report.html"), "generated_at": now,
    })
    run_log_path.write_text(json.dumps(run_log, indent=2), encoding="utf-8")
    base = results["base"]
    manifest = {
        "manifest_version": "1.0", "skill": "dcf-model-builder", "artifact_mode": "html_report_with_formula_workbook",
        "output_dir": str(directory), "model_status": "screen-grade",
        "first_read": {"path": str(directory / "report.html"), "role": "primary human deliverable", "why": "Open the standalone HTML report first for the conclusion, operating deep dive, valuation, and sources."},
        "primary_human_deliverable": str(directory / "report.html"),
        "human_deliverables": [{"path": str(directory / "report.html"), "role": "hero", "artifact_type": "html", "description": f"Standalone {CONTENT[company]['short_name']} intrinsic value report", "user_visible_default": True, "contains_new_analysis": True}],
        "companion_deliverables": [{"path": str(path), "role": "companion_workbook", "artifact_type": "xlsx", "description": "Formula-driven FCFF DCF workbook with Dashboard, scenarios, reverse DCF, sensitivities, checks, and sources", "user_visible_default": True, "contains_new_analysis": True}],
        "support_artifacts": [
            {"path": str(directory / "plan.json"), "role": "support_artifact", "artifact_type": "json", "description": "Normalized DCF plan and assumptions", "user_visible_default": False},
            {"path": str(directory / "model_citations.json"), "role": "support_artifact", "artifact_type": "json", "description": "Workbook cell citation ledger", "user_visible_default": False},
            {"path": str(run_log_path), "role": "support_artifact", "artifact_type": "json", "description": "Calculation, formula-error, and tie-out run log", "user_visible_default": False},
        ],
        "valuation_summary": {"current_price": plan["market"]["current_share_price"], "downside": results["downside"]["value_per_share"], "base": base["value_per_share"], "upside": results["upside"]["value_per_share"], "base_discount_to_market": base["value_per_share"] / plan["market"]["current_share_price"] - 1},
        "calculation_integrity": {"status": "pass" if not errors else "review", "formula_count": formula_count, "error_count": len(errors)},
        "decision_readiness": {"status": "screen-grade", "reason": "Forecast, WACC, and terminal assumptions are independent analyst estimates rather than management or consensus forecasts."},
        "generated_at": now,
    }
    (directory / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("stage", choices=("prepare", "finalize", "all"))
    args = parser.parse_args()
    if args.stage in ("prepare", "all"):
        for company in COMPANIES:
            prepare_workbook(company)
    if args.stage in ("finalize", "all"):
        for company in COMPANIES:
            finalize(company)


if __name__ == "__main__":
    main()
